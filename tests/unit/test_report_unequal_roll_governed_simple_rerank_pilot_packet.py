from __future__ import annotations

import csv
import json

from openpyxl import load_workbook

from infra.scripts import report_unequal_roll_governed_simple_rerank_pilot_packet as report


def _case(**overrides):
    row = {
        "variant_key": "simple_value_tier_rerank",
        "county_id": "harris",
        "subject_account": "A1",
        "neighborhood_code": "100.00",
        "subject_parcel_id": "subject-1",
        "governance_view": "automated_safe",
        "governance_classification": "eligible_candidate",
        "governance_reasons": "model_backed_stable_material_benefit",
        "model_backed": "True",
        "governed_taxpayer_delta_vs_similarity_top_100": "2500",
        "true_final_status_downgrade_raw": "False",
        "true_transition_to_unsupported_raw": "False",
        "included_comp_collapse_raw": "False",
        "rerank_replay_value_interpretation": "final_model_value",
        "final_status_transition": "supported_with_review -> supported_with_review",
        "smart_included_comp_count": "20",
        "rerank_final_included_comp_count": "20",
        "smart_vs_rerank_overlap_count": "10",
        "smart_full_included_comp_ids": "comp-1",
        "rerank_full_included_comp_ids": "comp-1",
        "added_comp_ids": "",
        "removed_comp_ids": "",
        "smart_value_interpretation": "final_model_value",
        "smart_final_value_status": "supported_with_review",
        "smart_requested_roll_value": "290000",
        "smart_requested_reduction_amount": "10000",
        "rerank_requested_roll_value": "287500",
        "rerank_requested_reduction_amount": "12500",
    }
    row.update(overrides)
    return row


def _comp(**overrides):
    row = {
        "variant_key": "simple_value_tier_rerank",
        "county_id": "harris",
        "subject_account": "A1",
        "neighborhood_code": "100.00",
        "comp_parcel_id": "comp-1",
        "comp_account_number": "C1",
        "comp_tax_year": "2025",
        "membership": "overlap",
        "total_abs_adjustment": "1200.0",
        "adjusted_value": "",
        "line_item_count": "10",
        "living_area_adjustment": "",
        "age_effective_age_adjustment": "",
        "full_bath_adjustment": "",
        "half_bath_adjustment": "",
        "land_site_adjustment": "",
        "story_adjustment": "",
        "pool_adjustment": "",
        "quality_adjustment": "",
        "condition_adjustment": "",
        "total_adjustment_amount": "",
        "adjustment_percent": "",
        "adjustment_source_status": "",
        "adjustment_line_items_json": "",
    }
    row.update(overrides)
    return row


def _write_json(path, payload):
    path.write_text(json.dumps(payload))


def _read_csv(path):
    with path.open(newline="") as fh:
        return list(csv.DictReader(fh))


def _first_row(rows, **matches):
    for row in rows:
        if all(row.get(key) == value for key, value in matches.items()):
            return row
    raise AssertionError(f"Missing row matching {matches}")


def _patch_hydration(monkeypatch):
    def fake_subjects(rows, database_url, requested_tax_year):
        return {
            row["subject_parcel_id"]: {
                "subject_parcel_id": row["subject_parcel_id"],
                "subject_address": f"{row['subject_account']} Main St",
                "subject_appraised_value": 300000,
                "subject_living_area_sf": 2000,
                "subject_value_per_sf": 150.0,
                "subject_year_built": 1990,
                "subject_land_sf": 8000,
                "subject_land_acres": 0.18,
                "subject_bedrooms": 3,
                "subject_full_baths": 2.0,
                "subject_half_baths": 1.0,
                "subject_stories": 1.0,
                "subject_subdivision": "TEST SUBDIVISION",
                "subject_neighborhood_code": row["neighborhood_code"],
                "subject_quality_code": "C",
                "subject_condition_code": "AVERAGE",
            }
            for row in rows
        }

    def fake_comps(rows, database_url, requested_tax_year):
        return {
            row["comp_parcel_id"]: {
                "comp_parcel_id": row["comp_parcel_id"],
                "comp_tax_year": requested_tax_year,
                "comp_account_number": row.get("comp_account_number"),
                "comp_address": f"{row.get('comp_account_number')} Oak Dr",
                "comp_appraised_value": 280000,
                "comp_living_area_sf": 1900,
                "comp_value_per_sf": 147.37,
                "comp_neighborhood_code": row["neighborhood_code"],
                "comp_subdivision_name": "TEST SUBDIVISION",
                "comp_land_sf": 7900,
                "comp_land_acres": 0.18,
                "comp_bedrooms": 3,
                "comp_full_baths": 2.0,
                "comp_half_baths": 1.0,
                "comp_stories": 1.0,
                "comp_year_built": 1991,
                "comp_quality_code": "C",
                "comp_condition_code": "AVERAGE",
                "line_item_count": row.get("line_item_count"),
            }
            for row in rows
        }

    monkeypatch.setattr(report, "hydrate_subject_facts", fake_subjects)
    monkeypatch.setattr(report, "hydrate_comp_facts", fake_comps)


def test_build_packet_splits_queues_and_corrects_tax_year(tmp_path, monkeypatch):
    _patch_hydration(monkeypatch)
    evidence_path = tmp_path / "evidence.json"
    fallback_path = tmp_path / "fallback.json"
    _write_json(
        evidence_path,
        {
            "case_rows": [
                _case(subject_account="A1", subject_parcel_id="subject-1", smart_vs_rerank_overlap_count="10"),
                _case(subject_account="A2", subject_parcel_id="subject-2", smart_vs_rerank_overlap_count="8"),
                _case(subject_account="A3", subject_parcel_id="subject-3", smart_vs_rerank_overlap_count="3"),
                _case(
                    county_id="fort_bend",
                    subject_account="F1",
                    neighborhood_code="200-00",
                    subject_parcel_id="subject-4",
                    governance_view="analyst_assisted",
                    governance_classification="manual_review_required",
                    true_final_status_downgrade_raw="True",
                ),
            ],
            "comp_rows": [
                _comp(subject_account="A1", comp_parcel_id="comp-1", comp_account_number="C1", membership="overlap"),
                _comp(subject_account="A2", comp_parcel_id="comp-2", comp_account_number="C2", membership="rerank_only"),
                _comp(subject_account="A3", comp_parcel_id="comp-3", comp_account_number="C3", membership="smart_only"),
                _comp(
                    county_id="fort_bend",
                    subject_account="F1",
                    neighborhood_code="200-00",
                    comp_parcel_id="comp-4",
                    comp_account_number="FC1",
                    membership="overlap",
                ),
            ],
        },
    )
    _write_json(
        fallback_path,
        {
            "case_rows": [
                {
                    "variant_key": "simple_value_tier_rerank",
                    "county_id": "harris",
                    "subject_account": "FB1",
                    "neighborhood_code": "100.00",
                    "governance_view": "fallback_blocked",
                    "fallback_to_similarity_top_100": True,
                    "governed_taxpayer_delta_vs_similarity_top_100": 0.0,
                }
            ]
        },
    )

    payload, paths = report.build_packet(
        complete_comp_evidence_artifact=evidence_path,
        governed_fallback_artifacts=[fallback_path],
        raw_artifacts=[],
        database_url="unused",
        requested_tax_year=2026,
        output_dir=tmp_path,
        timestamp="20260101T000000",
    )

    assert payload["summary"]["decision_counts"] == {
        "governed_rerank_ready": 1,
        "spot_check_only": 1,
        "analyst_review_only": 1,
        "hold_out": 1,
    }
    assert payload["summary"]["first_pilot_ready"]["case_count"] == 1
    assert payload["summary"]["governed_rerank_ready"]["case_count"] == 1
    assert payload["summary"]["fallback_blocked"]["case_count"] == 1
    assert payload["guardrails"]["db_writes"] is False

    comp_rows = _read_csv(paths["first_pilot_comp_details"])
    assert len(comp_rows) == 1
    assert comp_rows[0]["original_reported_comp_tax_year"] == "2025"
    assert comp_rows[0]["comp_tax_year"] == "2026"
    assert comp_rows[0]["membership"] == "overlap"
    assert paths["column_key"].exists()
    assert paths["pilot_summary"].exists()
    assert paths["qa_provenance"].exists()
    assert paths["comparison_grid"].exists()
    assert paths["changed_comps_review"].exists()
    assert paths["opinion_of_value"].exists()
    assert paths["governed_rerank_ready"].exists()
    assert paths["workbook"].exists()
    sheetnames = load_workbook(paths["workbook"], read_only=True).sheetnames
    assert "Executive_Summary" in sheetnames
    assert "QA_Provenance" in sheetnames
    assert "Opinion_Of_Value" in sheetnames
    assert "Governed_Rerank_Ready" in sheetnames


def test_subject_facts_are_repeated_in_signoff_and_comp_rows(tmp_path, monkeypatch):
    _patch_hydration(monkeypatch)
    evidence_path = tmp_path / "evidence.json"
    _write_json(
        evidence_path,
        {
            "case_rows": [
                _case(
                    county_id="fort_bend",
                    subject_account="F1",
                    neighborhood_code="200-00",
                    subject_parcel_id="subject-fb",
                )
            ],
            "comp_rows": [
                _comp(
                    county_id="fort_bend",
                    subject_account="F1",
                    neighborhood_code="200-00",
                    comp_parcel_id="comp-fb",
                    comp_account_number="FC1",
                    membership="rerank_only",
                )
            ],
        },
    )

    _, paths = report.build_packet(
        complete_comp_evidence_artifact=evidence_path,
        governed_fallback_artifacts=[],
        raw_artifacts=[],
        database_url="unused",
        requested_tax_year=2026,
        output_dir=tmp_path,
        timestamp="20260101T000001",
    )

    signoff = _read_csv(paths["signoff_tracker"])
    comp_rows = _read_csv(paths["first_pilot_comp_details"])

    assert signoff[0]["subject_address"] == "F1 Main St"
    assert signoff[0]["subject_full_baths"] == "2.0"
    assert comp_rows[0]["subject_address"] == "F1 Main St"
    assert comp_rows[0]["subject_full_baths"] == "2.0"
    assert comp_rows[0]["comp_full_baths"] == "2.0"
    assert comp_rows[0]["membership"] == "rerank_only"
    assert "adjustment burden" in comp_rows[0]["adjustment_burden_explanation"]


def test_simplified_signoff_and_review_surfaces_are_emitted(tmp_path, monkeypatch):
    _patch_hydration(monkeypatch)
    evidence_path = tmp_path / "evidence.json"
    _write_json(
        evidence_path,
        {
            "case_rows": [_case(added_comp_ids="comp-2", removed_comp_ids="comp-3")],
            "comp_rows": [
                _comp(comp_parcel_id="comp-1", comp_account_number="C1", membership="overlap"),
                _comp(comp_parcel_id="comp-2", comp_account_number="C2", membership="rerank_only"),
                _comp(comp_parcel_id="comp-3", comp_account_number="C3", membership="smart_only"),
            ],
        },
    )

    _, paths = report.build_packet(
        complete_comp_evidence_artifact=evidence_path,
        governed_fallback_artifacts=[],
        raw_artifacts=[],
        database_url="unused",
        requested_tax_year=2026,
        output_dir=tmp_path,
        timestamp="20260101T000004",
    )

    signoff = _read_csv(paths["signoff_tracker"])
    changed = _read_csv(paths["changed_comps_review"])
    grid = _read_csv(paths["comparison_grid"])
    key = _read_csv(paths["column_key"])
    opinion = _read_csv(paths["opinion_of_value"])

    assert list(signoff[0]) == report.SIGNOFF_FIELDNAMES
    assert {row["membership"] for row in changed} == {"rerank_only", "smart_only"}
    assert any(row["row_label"] == "Tax year" and row["SUBJECT"] == "2026" for row in grid)
    assert any(row["row_label"] == "Adjusted Appraised Value/SF" for row in grid)
    assert any(row["row_label"] == "Living area adjustment" for row in grid)
    assert any(
        row["row_label"] == "Line-item adjustment detail"
        and "unavailable_in_source_artifact" in row.values()
        for row in grid
    )
    assert any(row["row_label"] == "Line Item Count" for row in grid)
    assert any(row["field"] == "membership" for row in key)
    assert any(row["field"] == "opinion_of_value" for row in key)
    assert any(row["field"] == "final_value_formula" for row in key)
    conclusion = _first_row(opinion, row_type="conclusion")
    comp_row = _first_row(opinion, row_type="comp")
    assert conclusion["opinion_of_value"] == "287500.0"
    assert conclusion["model_opinion_value"] == "287500.0"
    assert conclusion["final_value_formula"] == "median_of_adjusted_appraised_values"
    assert conclusion["raw_psf_diagnostic_only_flag"] == "True"
    assert conclusion["weighted_psf_shortcut_used_flag"] == "False"
    assert conclusion["median_appraised_value_per_sf"] == "147.37"
    assert conclusion["adjusted_median_value_per_sf"] == ""
    assert comp_row["comp_adjusted_value_per_sf"] == ""
    assert comp_row["opinion_of_value"] == ""


def test_adjusted_value_per_sf_is_computed_only_when_supported(tmp_path, monkeypatch):
    _patch_hydration(monkeypatch)
    evidence_path = tmp_path / "evidence.json"
    _write_json(
        evidence_path,
        {
            "case_rows": [_case(rerank_requested_roll_value="295000", rerank_requested_reduction_amount="5000")],
            "comp_rows": [
                _comp(
                    adjusted_value="285000",
                    comp_parcel_id="comp-adjusted",
                    comp_account_number="C-adjusted",
                )
            ],
        },
    )

    _, paths = report.build_packet(
        complete_comp_evidence_artifact=evidence_path,
        governed_fallback_artifacts=[],
        raw_artifacts=[],
        database_url="unused",
        requested_tax_year=2026,
        output_dir=tmp_path,
        timestamp="20260101T000005",
    )

    opinion = _read_csv(paths["opinion_of_value"])
    grid = _read_csv(paths["comparison_grid"])

    comp_row = _first_row(opinion, row_type="comp")
    conclusion = _first_row(opinion, row_type="conclusion")
    assert comp_row["comp_adjusted_value_per_sf"] == "150.0"
    assert conclusion["adjusted_median_value_per_sf"] == "150.0"
    assert conclusion["median_adjusted_appraised_value"] == "285000.0"
    assert conclusion["psf_cross_check_value"] == "300000.0"
    assert conclusion["psf_cross_check_difference"] == "-5000.0"
    assert conclusion["psf_cross_check_difference_pct"] == "-0.0169"
    assert any(
        row["row_label"] == "Adjusted Appraised Value/SF" and row["OVERLAP COMP 1"] == "150.0"
        for row in grid
    )


def test_wrong_tax_year_after_hydration_keeps_case_out_of_first_pilot(tmp_path, monkeypatch):
    def fake_subjects(rows, database_url, requested_tax_year):
        return {
            row["subject_parcel_id"]: {
                "subject_parcel_id": row["subject_parcel_id"],
                "subject_address": "A1 Main St",
                "subject_appraised_value": 300000,
            }
            for row in rows
        }

    def fake_comps(rows, database_url, requested_tax_year):
        return {
            row["comp_parcel_id"]: {
                "comp_parcel_id": row["comp_parcel_id"],
                "comp_tax_year": 2025,
                "comp_account_number": row.get("comp_account_number"),
            }
            for row in rows
        }

    monkeypatch.setattr(report, "hydrate_subject_facts", fake_subjects)
    monkeypatch.setattr(report, "hydrate_comp_facts", fake_comps)
    evidence_path = tmp_path / "evidence.json"
    _write_json(evidence_path, {"case_rows": [_case()], "comp_rows": [_comp(comp_tax_year="2025")]})

    payload, _ = report.build_packet(
        complete_comp_evidence_artifact=evidence_path,
        governed_fallback_artifacts=[],
        raw_artifacts=[],
        database_url="unused",
        requested_tax_year=2026,
        output_dir=tmp_path,
        timestamp="20260101T000002",
    )

    assert payload["summary"]["governed_rerank_ready"]["case_count"] == 0
    assert payload["summary"]["analyst_review_only"]["case_count"] == 1
    assert "wrong_tax_year_comp_rows" in payload["analyst_review_rows"][0]["final_decision_reasons"]


def test_baseline_support_only_when_rerank_has_low_incremental_benefit(tmp_path, monkeypatch):
    _patch_hydration(monkeypatch)
    evidence_path = tmp_path / "evidence.json"
    _write_json(
        evidence_path,
        {
            "case_rows": [
                _case(
                    governance_view="fallback_blocked",
                    governance_classification="not_eligible_low_benefit",
                    governed_taxpayer_delta_vs_similarity_top_100="0",
                    smart_requested_roll_value="275000",
                    smart_requested_reduction_amount="25000",
                    rerank_requested_roll_value="275000",
                    rerank_requested_reduction_amount="25000",
                )
            ],
            "comp_rows": [_comp(membership="overlap")],
        },
    )

    payload, paths = report.build_packet(
        complete_comp_evidence_artifact=evidence_path,
        governed_fallback_artifacts=[],
        raw_artifacts=[],
        database_url="unused",
        requested_tax_year=2026,
        output_dir=tmp_path,
        timestamp="20260101T000006",
    )

    assert payload["summary"]["baseline_support_only"]["case_count"] == 1
    assert payload["summary"]["baseline_support_only"]["model_backed_net_savings"] == 25000
    assert payload["baseline_support_rows"][0]["final_decision"] == "baseline_support_only"
    assert payload["baseline_support_rows"][0]["packet_value_reduction_amount"] == 25000
    assert _read_csv(paths["baseline_support"])[0]["final_decision"] == "baseline_support_only"
    assert _read_csv(paths["baseline_support_comp_details"])[0]["membership"] == "overlap"
    opinion = _read_csv(paths["opinion_of_value"])
    conclusion = _first_row(opinion, row_type="conclusion")
    assert conclusion["opinion_source"] == "similarity_top_100_baseline"
    assert conclusion["opinion_of_value"] == "275000.0"
    assert conclusion["model_opinion_value"] == "275000.0"
    assert conclusion["final_value_formula"] == "median_of_adjusted_appraised_values"


def test_safety_blocked_rerank_can_use_safe_baseline_support(tmp_path, monkeypatch):
    _patch_hydration(monkeypatch)
    evidence_path = tmp_path / "evidence.json"
    _write_json(
        evidence_path,
        {
            "case_rows": [
                _case(
                    governance_view="fallback_blocked",
                    governance_classification="blocked_case",
                    true_transition_to_unsupported_raw="True",
                    governed_taxpayer_delta_vs_similarity_top_100="0",
                    smart_requested_reduction_amount="5000",
                )
            ],
            "comp_rows": [_comp(membership="overlap")],
        },
    )

    payload, _ = report.build_packet(
        complete_comp_evidence_artifact=evidence_path,
        governed_fallback_artifacts=[],
        raw_artifacts=[],
        database_url="unused",
        requested_tax_year=2026,
        output_dir=tmp_path,
        timestamp="20260101T000007",
    )

    assert payload["summary"]["baseline_support_only"]["case_count"] == 1
    assert payload["summary"]["fallback_safety_blocked"]["case_count"] == 0
    assert payload["summary"]["baseline_support_only"]["unsupported_transition_count"] == 0


def test_no_reduction_and_safety_blocked_modes_are_separated(tmp_path, monkeypatch):
    _patch_hydration(monkeypatch)
    evidence_path = tmp_path / "evidence.json"
    _write_json(
        evidence_path,
        {
            "case_rows": [
                _case(
                    subject_account="NORED",
                    subject_parcel_id="subject-nored",
                    governance_view="fallback_blocked",
                    governance_classification="not_eligible_low_benefit",
                    governed_taxpayer_delta_vs_similarity_top_100="0",
                    smart_requested_reduction_amount="0",
                ),
                _case(
                    subject_account="BLOCK",
                    subject_parcel_id="subject-block",
                    governance_view="fallback_blocked",
                    governance_classification="blocked_case",
                    governed_taxpayer_delta_vs_similarity_top_100="0",
                    smart_value_interpretation="diagnostic",
                    smart_requested_reduction_amount="10000",
                ),
            ],
            "comp_rows": [
                _comp(
                    subject_account="NORED",
                    comp_parcel_id="comp-nored",
                    living_area_adjustment="-1000",
                    age_effective_age_adjustment="500",
                    total_adjustment_amount="-500",
                    adjustment_percent="-0.0017",
                    adjustment_source_status="line_items_available",
                    adjustment_line_items_json='[{"adjustment_type": "gla", "signed_adjustment_amount": -1000}]',
                ),
                _comp(subject_account="BLOCK", comp_parcel_id="comp-block"),
            ],
        },
    )

    payload, paths = report.build_packet(
        complete_comp_evidence_artifact=evidence_path,
        governed_fallback_artifacts=[],
        raw_artifacts=[],
        database_url="unused",
        requested_tax_year=2026,
        output_dir=tmp_path,
        timestamp="20260101T000008",
    )

    assert payload["summary"]["no_reduction_no_action"]["case_count"] == 1
    assert payload["summary"]["fallback_safety_blocked"]["case_count"] == 1
    assert len(payload["no_action_review_rows"]) == 1
    assert len(payload["no_action_opinion_of_value_rows"]) == 3
    assert any(
        row["row_label"] == "Living area adjustment" and row["Selected Comp 1"] == "-1000"
        for row in payload["no_action_subject_comp_grid_rows"]
    )
    assert any(
        row["row_label"] == "Membership" and row["Selected Comp 1"] == "overlap"
        for row in payload["no_action_subject_comp_grid_rows"]
    )
    no_action_review = _read_csv(paths["no_action_review"])
    no_action_opinion = _read_csv(paths["no_action_opinion_of_value"])
    no_action_grid = _read_csv(paths["no_action_comparison_grid"])
    assert "Selected Comp 1" in no_action_grid[0]
    assert not any("RERANK" in fieldname for fieldname in no_action_grid[0])
    assert no_action_review[0]["packet_mode"] == "no_reduction_no_action"
    assert any(row["row_type"] == "conclusion" for row in no_action_opinion)
    assert any(row["row_label"] == "Line-item adjustment detail" for row in no_action_grid)


def test_guardrail_metadata_discloses_no_production_changes(tmp_path, monkeypatch):
    _patch_hydration(monkeypatch)
    evidence_path = tmp_path / "evidence.json"
    _write_json(evidence_path, {"case_rows": [_case()], "comp_rows": [_comp()]})

    payload, _ = report.build_packet(
        complete_comp_evidence_artifact=evidence_path,
        governed_fallback_artifacts=[],
        raw_artifacts=[],
        database_url="unused",
        requested_tax_year=2026,
        output_dir=tmp_path,
        timestamp="20260101T000003",
    )

    assert payload["artifact_contract"]["primary_variant"] == "simple_value_tier_rerank"
    assert payload["artifact_contract"]["bounded_proxy_used_for_conclusions"] is False
    assert payload["guardrails"]["db_writes"] is False
    assert payload["guardrails"]["migrations"] is False
    assert payload["guardrails"]["runtime_defaults_changed"] is False
    assert payload["guardrails"]["production_scoring_adjustment_median_governance_final_value_changed"] is False
    assert payload["guardrails"]["new_model_features_or_penalties"] is False
