from __future__ import annotations

import csv
import json

from openpyxl import load_workbook

from infra.scripts import report_unequal_roll_mvp_analyst_acceptance_tracker as report


def _packet_row(**overrides):
    row = {
        "county_id": "harris",
        "subject_account": "A1",
        "neighborhood_code": "100.00",
        "subject_address": "1 Main St",
        "subject_appraised_value": 300000,
        "subject_living_area_sf": 2000,
        "subject_value_per_sf": 150,
        "smart_requested_roll_value": 292000,
        "smart_requested_reduction_amount": 8000,
        "smart_included_comp_count": 20,
        "smart_value_interpretation": "final_model_value",
        "rerank_requested_roll_value": 287500,
        "rerank_requested_reduction_amount": 12500,
        "rerank_final_included_comp_count": 20,
        "rerank_value_interpretation": "final_model_value",
        "packet_value_reduction_amount": 12500,
        "final_decision": "governed_rerank_ready",
        "final_decision_reasons": ["model_backed", "material_gain"],
    }
    row.update(overrides)
    return row


def _packet(**overrides):
    payload = {
        "artifact_contract": {
            "artifact_type": "governed_simple_value_tier_rerank_mvp_pilot_packet",
            "architecture_name": "governed_similarity_baseline_with_simple_rerank",
            "created_at": "20260101T000000",
            "requested_tax_year": 2026,
            "bounded_proxy_used_for_conclusions": False,
            "final_requested_value_formula": "median_of_adjusted_appraised_values",
        },
        "case_rows": [
            _packet_row(),
            _packet_row(
                subject_account="B1",
                subject_address="2 Main St",
                final_decision="baseline_support_only",
                smart_requested_roll_value=290000,
                smart_requested_reduction_amount=10000,
                packet_value_reduction_amount=10000,
            ),
            _packet_row(
                subject_account="N1",
                subject_address="3 Main St",
                final_decision="no_reduction_no_action",
                smart_requested_roll_value=299500,
                smart_requested_reduction_amount=500,
                packet_value_reduction_amount=500,
            ),
        ],
    }
    payload.update(overrides)
    return payload


def _write_packet(path, payload=None):
    path.write_text(json.dumps(payload or _packet()))


def _read_csv(path):
    with path.open(newline="") as handle:
        return list(csv.DictReader(handle))


def test_build_tracker_extracts_core_queue_rows_and_counts(tmp_path):
    packet_path = tmp_path / "packet.json"
    _write_packet(packet_path)

    tracker = report.build_tracker([packet_path], label="pilot", analyst_name="Analyst A", review_date="2026-05-15")

    rows = tracker["review_queue_rows"]
    assert [row["packet_mode"] for row in rows] == [
        "governed_rerank_ready",
        "baseline_support_only",
        "no_reduction_no_action",
    ]
    assert rows[0]["model_value_source"] == "simple_value_tier_rerank"
    assert rows[0]["model_supported_value"] == 287500
    assert rows[0]["reduction_amount"] == 12500
    assert rows[1]["model_value_source"] == "similarity_top_100_baseline"
    assert rows[1]["model_supported_value"] == 290000
    assert rows[2]["model_value_source"] == "similarity_top_100_no_action"
    assert rows[0]["reviewed_by"] == "Analyst A"

    by_queue = {row["packet_mode"]: row for row in tracker["summary"]["by_queue"]}
    assert by_queue["governed_rerank_ready"]["case_count"] == 1
    assert by_queue["baseline_support_only"]["case_count"] == 1
    assert by_queue["no_reduction_no_action"]["case_count"] == 1
    assert tracker["summary"]["pending_review_count"] == 3


def test_write_outputs_emits_workbook_keys_and_non_raw_review_queue(tmp_path):
    packet_path = tmp_path / "packet.json"
    _write_packet(packet_path)
    tracker = report.build_tracker([packet_path], label="pilot")

    paths = report.write_outputs(tracker, tmp_path, "pilot")

    workbook = load_workbook(paths["workbook"], read_only=False)
    assert {
        "README",
        "Review_Queue",
        "Comp_Issue_Log",
        "Summary",
        "Decision_Key",
        "Issue_Key",
        "Raw_Input_Index",
    }.issubset(set(workbook.sheetnames))

    review_headers = [cell.value for cell in workbook["Review_Queue"][1]]
    assert review_headers == report.REVIEW_QUEUE_FIELDNAMES
    assert "governance_reasons" not in review_headers
    assert "smart_full_included_comp_ids" not in review_headers
    assert "rerank_full_included_comp_ids" not in review_headers
    assert "analyst_decision" in review_headers
    assert workbook["Review_Queue"].auto_filter.ref is not None
    assert workbook["Review_Queue"].freeze_panes == "A2"
    assert workbook["Comp_Issue_Log"].max_column == len(report.COMP_ISSUE_FIELDNAMES)

    decision_rows = list(workbook["Decision_Key"].iter_rows(min_row=2, values_only=True))
    issue_rows = list(workbook["Issue_Key"].iter_rows(min_row=2, values_only=True))
    assert any(row[0] == "approve" for row in decision_rows)
    assert any(row[0] == "comp_location_issue" for row in issue_rows)

    csv_rows = _read_csv(paths["review_queue_csv"])
    assert len(csv_rows) == 3
    assert csv_rows[0]["packet_label"] == "pilot"


def test_source_artifact_provenance_is_preserved(tmp_path):
    packet_path = tmp_path / "packet.json"
    _write_packet(packet_path)
    tracker = report.build_tracker([packet_path], label="pilot")

    raw_index = tracker["raw_input_index_rows"]
    assert len(raw_index) == 3
    assert raw_index[0]["source_packet_artifact"] == str(packet_path)
    assert raw_index[0]["source_account"] == "A1"
    assert raw_index[0]["requested_tax_year"] == 2026
    assert raw_index[0]["bounded_proxy_used_for_conclusions"] is False


def test_comma_separated_packet_arguments_are_supported(tmp_path):
    first = tmp_path / "first.json"
    second = tmp_path / "second.json"
    _write_packet(first, _packet(case_rows=[_packet_row(subject_account="A1")]))
    _write_packet(second, _packet(case_rows=[_packet_row(subject_account="A2")]))

    paths = report._split_artifact_args([f"{first},{second}"])

    assert paths == [first, second]
