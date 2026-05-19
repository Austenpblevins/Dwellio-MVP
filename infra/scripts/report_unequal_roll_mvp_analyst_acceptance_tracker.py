#!/usr/bin/env python3
"""Build analyst acceptance trackers from unequal-roll MVP packet artifacts.

This report is downstream of packet generation. It does not rerun candidate
selection, reranking, adjustment math, final-value calculation, or governance.
It creates an analyst-friendly workbook and summary artifacts for recording
whether packet outputs are acceptable for MVP rollout decisions.
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover - optional workbook output
    Workbook = None
    Alignment = Font = PatternFill = get_column_letter = DataValidation = None


ARCHITECTURE_NAME = "governed_similarity_baseline_with_simple_rerank"
FINAL_VALUE_FORMULA = "median_of_adjusted_appraised_values"

PACKET_MODE_ORDER = [
    "governed_rerank_ready",
    "baseline_support_only",
    "spot_check_only",
    "analyst_review_only",
    "hold_out",
    "fallback_safety_blocked",
    "no_reduction_no_action",
]

DECISION_CHOICES = [
    "approve",
    "approve_with_note",
    "reject_comp_quality",
    "reject_subject_data",
    "reject_adjustment_issue",
    "reject_too_aggressive",
    "reject_too_conservative",
    "needs_more_review",
    "hold_out",
]

REJECTION_REASON_CHOICES = [
    "comp_location_issue",
    "comp_size_issue",
    "comp_age_issue",
    "comp_land_issue",
    "comp_bath_bed_issue",
    "comp_quality_condition_issue",
    "adjustment_amount_issue",
    "subject_data_issue",
    "missing_evidence_issue",
    "packet_clarity_issue",
    "too_aggressive",
    "too_conservative",
    "other",
]

ISSUE_FLAG_CHOICES = [
    "comp_location_issue",
    "comp_size_issue",
    "comp_age_issue",
    "comp_land_issue",
    "comp_bath_bed_issue",
    "comp_quality_condition_issue",
    "adjustment_amount_issue",
    "subject_data_issue",
    "missing_evidence_issue",
    "packet_clarity_issue",
]

REVIEW_QUEUE_FIELDNAMES = [
    "tracker_row_id",
    "packet_label",
    "packet_mode",
    "county",
    "neighborhood",
    "account",
    "property_address",
    "current_appraised_value",
    "model_supported_value",
    "reduction_amount",
    "reduction_percent",
    "final_value_formula",
    "selected_comp_count",
    "model_value_source",
    "final_value_status",
    "reason_summary",
    "analyst_decision",
    "rejection_reason",
    "issue_flags",
    "notes",
    "reviewed_by",
    "reviewed_at",
]

COMP_ISSUE_FIELDNAMES = [
    "account",
    "comp_account",
    "comp_membership",
    "issue_type",
    "issue_description",
    "analyst_action",
    "notes",
]

RAW_INPUT_FIELDNAMES = [
    "tracker_row_id",
    "source_packet_artifact",
    "packet_label",
    "source_packet_created_at",
    "architecture_name",
    "requested_tax_year",
    "bounded_proxy_used_for_conclusions",
    "source_queue",
    "packet_mode",
    "source_account",
    "county",
    "neighborhood",
    "source_row_index",
]

SUMMARY_QUEUE_FIELDNAMES = [
    "packet_mode",
    "case_count",
    "model_proposed_value_reduction",
    "pending_review_count",
]

SUMMARY_COUNTY_FIELDNAMES = [
    "county",
    "case_count",
    "model_proposed_value_reduction",
    "pending_review_count",
]

SUMMARY_SEGMENT_FIELDNAMES = [
    "county",
    "neighborhood",
    "case_count",
    "model_proposed_value_reduction",
    "pending_review_count",
]

QUEUE_FILLS = {
    "governed_rerank_ready": "D9EAD3",
    "baseline_support_only": "DDEBF7",
    "spot_check_only": "FFF2CC",
    "analyst_review_only": "FCE4D6",
    "hold_out": "E7E6E6",
    "fallback_safety_blocked": "F4CCCC",
    "no_reduction_no_action": "EADCF8",
}


def _as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _round_money(value: Any) -> float | None:
    numeric = _as_float(value)
    if numeric is None:
        return None
    return round(numeric, 2)


def _format_money(value: Any) -> str:
    numeric = _as_float(value)
    if numeric is None:
        return ""
    return f"${numeric:,.2f}"


def _packet_mode(row: dict[str, Any]) -> str:
    mode = row.get("final_decision") or row.get("packet_mode") or row.get("queue")
    return str(mode or "unknown")


def _reason_summary(row: dict[str, Any]) -> str:
    reasons = row.get("final_decision_reasons") or row.get("governance_reasons") or row.get("fallback_reasons")
    if isinstance(reasons, list):
        return "; ".join(str(reason) for reason in reasons if reason not in (None, ""))
    if reasons in (None, ""):
        return ""
    return str(reasons)


def _model_source_for_mode(mode: str) -> str:
    if mode == "governed_rerank_ready":
        return "simple_value_tier_rerank"
    if mode == "baseline_support_only":
        return "similarity_top_100_baseline"
    if mode == "no_reduction_no_action":
        return "similarity_top_100_no_action"
    if mode == "fallback_safety_blocked":
        return "safety_blocked"
    if mode in {"spot_check_only", "analyst_review_only", "hold_out"}:
        return mode
    return mode or "unknown"


def _supported_value(row: dict[str, Any], mode: str) -> float | None:
    if mode == "governed_rerank_ready":
        return _round_money(
            row.get("rerank_requested_roll_value")
            or row.get("model_supported_value")
            or row.get("packet_supported_value")
        )
    if mode in {"baseline_support_only", "no_reduction_no_action", "fallback_safety_blocked"}:
        return _round_money(
            row.get("smart_requested_roll_value")
            or row.get("similarity_top_100_value")
            or row.get("model_supported_value")
        )
    return _round_money(
        row.get("model_supported_value")
        or row.get("packet_supported_value")
        or row.get("rerank_requested_roll_value")
        or row.get("smart_requested_roll_value")
    )


def _reduction_amount(row: dict[str, Any], mode: str, supported_value: float | None) -> float | None:
    explicit = (
        row.get("packet_value_reduction_amount")
        or row.get("governed_taxpayer_savings")
        or row.get("baseline_value_reduction_amount")
    )
    if mode == "baseline_support_only":
        explicit = row.get("baseline_value_reduction_amount") or row.get("smart_requested_reduction_amount") or explicit
    elif mode == "no_reduction_no_action":
        explicit = row.get("smart_requested_reduction_amount") or row.get("baseline_value_reduction_amount") or explicit
    elif mode == "governed_rerank_ready":
        explicit = row.get("packet_value_reduction_amount") or row.get("rerank_requested_reduction_amount") or explicit

    numeric = _round_money(explicit)
    if numeric is not None:
        return numeric

    appraised = _as_float(row.get("subject_appraised_value") or row.get("current_appraised_value"))
    if appraised is None or supported_value is None:
        return None
    return round(max(0.0, appraised - supported_value), 2)


def _selected_comp_count(row: dict[str, Any], mode: str) -> int | None:
    if mode == "governed_rerank_ready":
        count = row.get("rerank_final_included_comp_count") or row.get("rerank_included_comp_count")
    else:
        count = row.get("smart_final_included_comp_count") or row.get("smart_included_comp_count")
    if count in (None, ""):
        count = row.get("selected_comp_count") or row.get("included_comp_count")
    try:
        return int(count)
    except (TypeError, ValueError):
        return None


def _final_value_status(row: dict[str, Any], mode: str) -> str:
    if mode == "governed_rerank_ready":
        return str(row.get("rerank_value_interpretation") or row.get("rerank_final_status") or "")
    return str(row.get("smart_value_interpretation") or row.get("smart_final_status") or row.get("value_interpretation") or "")


def _artifact_label(path: Path, artifact: dict[str, Any], label: str | None, index: int) -> str:
    if label:
        return label if index == 0 else f"{label}_{index + 1}"
    contract = artifact.get("artifact_contract") or {}
    created = contract.get("created_at")
    if created:
        return str(created)
    return path.stem


def _split_artifact_args(values: list[str]) -> list[Path]:
    paths: list[Path] = []
    for value in values:
        for raw_part in value.split(","):
            part = raw_part.strip()
            if part:
                paths.append(Path(part))
    return paths


def build_tracker(
    packet_artifact_paths: list[Path],
    *,
    label: str | None = None,
    analyst_name: str | None = None,
    review_date: str | None = None,
) -> dict[str, Any]:
    review_rows: list[dict[str, Any]] = []
    raw_index_rows: list[dict[str, Any]] = []
    source_summaries: list[dict[str, Any]] = []
    tracker_row_id = 1

    for artifact_index, packet_path in enumerate(packet_artifact_paths):
        artifact = json.loads(packet_path.read_text())
        contract = artifact.get("artifact_contract") or {}
        packet_label = _artifact_label(packet_path, artifact, label, artifact_index)
        case_rows = artifact.get("case_rows") or []
        source_summaries.append(
            {
                "packet_label": packet_label,
                "source_packet_artifact": str(packet_path),
                "case_count": len(case_rows),
                "created_at": contract.get("created_at"),
                "architecture_name": contract.get("architecture_name") or contract.get("architecture"),
                "requested_tax_year": contract.get("requested_tax_year"),
                "bounded_proxy_used_for_conclusions": contract.get("bounded_proxy_used_for_conclusions"),
            }
        )

        for source_row_index, row in enumerate(case_rows, start=1):
            mode = _packet_mode(row)
            supported_value = _supported_value(row, mode)
            reduction = _reduction_amount(row, mode, supported_value)
            appraised = _as_float(row.get("subject_appraised_value") or row.get("current_appraised_value"))
            reduction_percent = round(reduction / appraised, 6) if appraised and reduction is not None else None
            review_row = {
                "tracker_row_id": tracker_row_id,
                "packet_label": packet_label,
                "packet_mode": mode,
                "county": row.get("county_id") or row.get("county"),
                "neighborhood": row.get("neighborhood_code") or row.get("subject_neighborhood_code"),
                "account": row.get("subject_account") or row.get("account"),
                "property_address": row.get("subject_address") or row.get("property_address"),
                "current_appraised_value": _round_money(row.get("subject_appraised_value") or row.get("current_appraised_value")),
                "model_supported_value": supported_value,
                "reduction_amount": reduction,
                "reduction_percent": reduction_percent,
                "final_value_formula": contract.get("final_requested_value_formula") or FINAL_VALUE_FORMULA,
                "selected_comp_count": _selected_comp_count(row, mode),
                "model_value_source": _model_source_for_mode(mode),
                "final_value_status": _final_value_status(row, mode),
                "reason_summary": _reason_summary(row),
                "analyst_decision": "",
                "rejection_reason": "",
                "issue_flags": "",
                "notes": "",
                "reviewed_by": analyst_name or "",
                "reviewed_at": review_date or "",
            }
            review_rows.append(review_row)

            raw_index_rows.append(
                {
                    "tracker_row_id": tracker_row_id,
                    "source_packet_artifact": str(packet_path),
                    "packet_label": packet_label,
                    "source_packet_created_at": contract.get("created_at"),
                    "architecture_name": contract.get("architecture_name") or contract.get("architecture"),
                    "requested_tax_year": contract.get("requested_tax_year"),
                    "bounded_proxy_used_for_conclusions": contract.get("bounded_proxy_used_for_conclusions"),
                    "source_queue": mode,
                    "packet_mode": mode,
                    "source_account": row.get("subject_account") or row.get("account"),
                    "county": row.get("county_id") or row.get("county"),
                    "neighborhood": row.get("neighborhood_code") or row.get("subject_neighborhood_code"),
                    "source_row_index": source_row_index,
                }
            )
            tracker_row_id += 1

    summaries = summarize_review_rows(review_rows)
    return {
        "artifact_contract": {
            "artifact_type": "unequal_roll_mvp_analyst_acceptance_tracker",
            "architecture_name": ARCHITECTURE_NAME,
            "created_at": datetime.now().strftime("%Y%m%dT%H%M%S"),
            "input_artifact_count": len(packet_artifact_paths),
            "model_behavior_changed": False,
            "no_persist_reporting_only": True,
        },
        "source_packets": source_summaries,
        "summary": summaries,
        "review_queue_rows": review_rows,
        "comp_issue_log_template_rows": [],
        "decision_key_rows": decision_key_rows(),
        "issue_key_rows": issue_key_rows(),
        "raw_input_index_rows": raw_index_rows,
    }


def summarize_review_rows(review_rows: list[dict[str, Any]]) -> dict[str, Any]:
    queue_counts: Counter[str] = Counter()
    queue_values: defaultdict[str, float] = defaultdict(float)
    county_counts: Counter[str] = Counter()
    county_values: defaultdict[str, float] = defaultdict(float)
    segment_counts: Counter[tuple[str, str]] = Counter()
    segment_values: defaultdict[tuple[str, str], float] = defaultdict(float)

    for row in review_rows:
        mode = str(row.get("packet_mode") or "unknown")
        county = str(row.get("county") or "unknown")
        neighborhood = str(row.get("neighborhood") or "unknown")
        reduction = _as_float(row.get("reduction_amount")) or 0.0
        queue_counts[mode] += 1
        queue_values[mode] += reduction
        county_counts[county] += 1
        county_values[county] += reduction
        segment_counts[(county, neighborhood)] += 1
        segment_values[(county, neighborhood)] += reduction

    by_queue = [
        {
            "packet_mode": mode,
            "case_count": queue_counts.get(mode, 0),
            "model_proposed_value_reduction": round(queue_values.get(mode, 0.0), 2),
            "pending_review_count": queue_counts.get(mode, 0),
        }
        for mode in PACKET_MODE_ORDER
        if queue_counts.get(mode, 0)
    ]
    for mode in sorted(set(queue_counts) - set(PACKET_MODE_ORDER)):
        by_queue.append(
            {
                "packet_mode": mode,
                "case_count": queue_counts[mode],
                "model_proposed_value_reduction": round(queue_values[mode], 2),
                "pending_review_count": queue_counts[mode],
            }
        )

    by_county = [
        {
            "county": county,
            "case_count": county_counts[county],
            "model_proposed_value_reduction": round(county_values[county], 2),
            "pending_review_count": county_counts[county],
        }
        for county in sorted(county_counts)
    ]
    by_segment = [
        {
            "county": county,
            "neighborhood": neighborhood,
            "case_count": segment_counts[(county, neighborhood)],
            "model_proposed_value_reduction": round(segment_values[(county, neighborhood)], 2),
            "pending_review_count": segment_counts[(county, neighborhood)],
        }
        for county, neighborhood in sorted(segment_counts)
    ]

    total_value = round(sum((_as_float(row.get("reduction_amount")) or 0.0) for row in review_rows), 2)
    return {
        "total_review_rows": len(review_rows),
        "pending_review_count": len(review_rows),
        "model_proposed_value_reduction": total_value,
        "approval_rate": None,
        "rejection_rate": None,
        "acceptance_metrics_note": (
            "Approval/rejection rates are intentionally blank until analysts fill the tracker. "
            "Use the completed tracker to calculate post-review acceptance metrics."
        ),
        "by_queue": by_queue,
        "by_county": by_county,
        "by_segment": by_segment,
    }


def decision_key_rows() -> list[dict[str, str]]:
    definitions = {
        "approve": "Analyst accepts the model-supported value and comp evidence as pilot-usable.",
        "approve_with_note": "Analyst accepts the case, with a note for context or minor caveat.",
        "reject_comp_quality": "Comps are not credible enough for support.",
        "reject_subject_data": "Subject facts appear incorrect or incomplete.",
        "reject_adjustment_issue": "Adjustment amount, direction, or support appears unreliable.",
        "reject_too_aggressive": "Recommended value reduction is too aggressive for the evidence.",
        "reject_too_conservative": "Packet appears to leave material defensible value uncaptured.",
        "needs_more_review": "Case needs another analyst, data pull, or engineering review.",
        "hold_out": "Do not use this case in the first pilot.",
    }
    return [{"decision": key, "definition": definitions[key]} for key in DECISION_CHOICES]


def issue_key_rows() -> list[dict[str, str]]:
    definitions = {
        "comp_location_issue": "Comparable location, neighborhood, subdivision, or proximity concern.",
        "comp_size_issue": "Living area or scale mismatch concern.",
        "comp_age_issue": "Year built or effective age mismatch concern.",
        "comp_land_issue": "Land/site size or site utility mismatch concern.",
        "comp_bath_bed_issue": "Bedroom, full bath, half bath, or story mismatch concern.",
        "comp_quality_condition_issue": "Quality, condition, remodel, or physical comparability concern.",
        "adjustment_amount_issue": "Adjustment amount, direction, or burden concern.",
        "subject_data_issue": "Subject facts look wrong or incomplete.",
        "missing_evidence_issue": "Packet lacks facts needed for a decision.",
        "packet_clarity_issue": "Packet is hard to interpret or needs presentation cleanup.",
    }
    return [{"issue_flag": key, "definition": definitions[key]} for key in ISSUE_FLAG_CHOICES]


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _append_rows(ws: Any, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(field) for field in fieldnames])


def _format_sheet(ws: Any, *, currency_columns: set[str] | None = None, percent_columns: set[str] | None = None) -> None:
    currency_columns = currency_columns or set()
    percent_columns = percent_columns or set()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    if ws.max_column:
        ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    header_by_index = {index: ws.cell(row=1, column=index).value for index in range(1, ws.max_column + 1)}
    for column_index in range(1, ws.max_column + 1):
        header = str(header_by_index[column_index] or "")
        width = min(max(len(header) + 2, 12), 34)
        if header in {"reason_summary", "notes", "issue_description", "definition"}:
            width = 42
        ws.column_dimensions[get_column_letter(column_index)].width = width
        if header in currency_columns:
            for row_index in range(2, ws.max_row + 1):
                ws.cell(row=row_index, column=column_index).number_format = "$#,##0.00"
        if header in percent_columns:
            for row_index in range(2, ws.max_row + 1):
                ws.cell(row=row_index, column=column_index).number_format = "0.00%"

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _apply_review_validations(ws: Any) -> None:
    if ws.max_row < 2:
        return
    headers = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column + 1)]
    validation_specs = {
        "analyst_decision": DECISION_CHOICES,
        "rejection_reason": REJECTION_REASON_CHOICES,
        "issue_flags": ISSUE_FLAG_CHOICES,
    }
    for header, choices in validation_specs.items():
        if header not in headers:
            continue
        column_index = headers.index(header) + 1
        column_letter = get_column_letter(column_index)
        formula = '"' + ",".join(choices) + '"'
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Choose a value from the list, or leave blank until reviewed."
        validation.prompt = "Select a standard value. For multiple issue flags, add details in notes."
        ws.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{max(ws.max_row + 100, 200)}")


def _apply_queue_fills(ws: Any) -> None:
    headers = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column + 1)]
    if "packet_mode" not in headers:
        return
    mode_column = headers.index("packet_mode") + 1
    for row_index in range(2, ws.max_row + 1):
        mode = ws.cell(row=row_index, column=mode_column).value
        fill_color = QUEUE_FILLS.get(str(mode or ""))
        if not fill_color:
            continue
        fill = PatternFill("solid", fgColor=fill_color)
        for column_index in range(1, ws.max_column + 1):
            ws.cell(row=row_index, column=column_index).fill = fill


def write_workbook(path: Path, tracker: dict[str, Any]) -> None:
    if Workbook is None:  # pragma: no cover
        raise RuntimeError("openpyxl is required for workbook output")

    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    readme = workbook.create_sheet("README")
    readme_rows = [
        ("Purpose", "Track analyst acceptance decisions for no-persist unequal-roll MVP packet outputs."),
        ("How to use", "Review each row in Review_Queue, choose an analyst decision, fill issue flags/reason fields, and use Comp_Issue_Log for comp-specific concerns."),
        ("What not to edit", "Do not edit source provenance columns or packet-mode classifications; use analyst fields for review outcomes."),
        ("Multiple comp issues", "List specific comp accounts and reasons in Comp_Issue_Log; summarize subject-level impact in Review_Queue notes."),
        ("MVP boundary", "This tracker does not write production values and does not change model, scoring, adjustment, or governance logic."),
    ]
    readme.append(["section", "detail"])
    for row in readme_rows:
        readme.append(row)
    _format_sheet(readme)

    review = workbook.create_sheet("Review_Queue")
    _append_rows(review, REVIEW_QUEUE_FIELDNAMES, tracker["review_queue_rows"])
    _format_sheet(
        review,
        currency_columns={"current_appraised_value", "model_supported_value", "reduction_amount"},
        percent_columns={"reduction_percent"},
    )
    _apply_review_validations(review)
    _apply_queue_fills(review)

    comp_issue_log = workbook.create_sheet("Comp_Issue_Log")
    _append_rows(comp_issue_log, COMP_ISSUE_FIELDNAMES, [])
    _format_sheet(comp_issue_log)

    summary = workbook.create_sheet("Summary")
    summary.append(["section", "name", "case_count", "model_proposed_value_reduction", "pending_review_count"])
    summary.append([
        "overall",
        "all packets",
        tracker["summary"]["total_review_rows"],
        tracker["summary"]["model_proposed_value_reduction"],
        tracker["summary"]["pending_review_count"],
    ])
    for row in tracker["summary"]["by_queue"]:
        summary.append(["queue", row["packet_mode"], row["case_count"], row["model_proposed_value_reduction"], row["pending_review_count"]])
    for row in tracker["summary"]["by_county"]:
        summary.append(["county", row["county"], row["case_count"], row["model_proposed_value_reduction"], row["pending_review_count"]])
    _format_sheet(summary, currency_columns={"model_proposed_value_reduction"})

    decision_key = workbook.create_sheet("Decision_Key")
    _append_rows(decision_key, ["decision", "definition"], tracker["decision_key_rows"])
    _format_sheet(decision_key)

    issue_key = workbook.create_sheet("Issue_Key")
    _append_rows(issue_key, ["issue_flag", "definition"], tracker["issue_key_rows"])
    _format_sheet(issue_key)

    raw_input = workbook.create_sheet("Raw_Input_Index")
    _append_rows(raw_input, RAW_INPUT_FIELDNAMES, tracker["raw_input_index_rows"])
    _format_sheet(raw_input)

    workbook.save(path)


def write_markdown(path: Path, tracker: dict[str, Any]) -> None:
    lines = [
        "# Unequal-Roll MVP Analyst Acceptance Tracker",
        "",
        "This tracker is reporting-only. It summarizes packet outputs and provides a place for analysts to record acceptance decisions.",
        "",
        "## Source Packets",
    ]
    for source in tracker["source_packets"]:
        lines.append(f"- `{source['source_packet_artifact']}`: {source['case_count']} rows")

    lines.extend(
        [
            "",
            "## Coverage",
            f"- Review queue rows: {tracker['summary']['total_review_rows']}",
            f"- Pending review rows: {tracker['summary']['pending_review_count']}",
            f"- Model-proposed value reduction: {_format_money(tracker['summary']['model_proposed_value_reduction'])}",
            "",
            "## Counts By Packet Mode",
            "| Packet mode | Rows | Model-proposed value |",
            "| --- | ---: | ---: |",
        ]
    )
    for row in tracker["summary"]["by_queue"]:
        lines.append(
            f"| {row['packet_mode']} | {row['case_count']} | {_format_money(row['model_proposed_value_reduction'])} |"
        )

    lines.extend(
        [
            "",
            "## Acceptance Metrics",
            "Approval and rejection rates are intentionally pending until analysts fill the tracker. After review, calculate approval rates by queue, county, and neighborhood; rejection reason counts; issue flag counts; material value approved or rejected; and cases needing engineering follow-up.",
            "",
            "## Analyst Guidance",
            "- Use `Review_Queue` for subject-level decisions.",
            "- Use `Comp_Issue_Log` to list specific comp account issues.",
            "- Keep packet classifications intact; record disagreements in analyst fields.",
            "- Treat this as no-persist MVP acceptance evidence, not production value writing.",
        ]
    )
    path.write_text("\n".join(lines) + "\n")


def write_outputs(tracker: dict[str, Any], output_dir: Path, label: str | None) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = tracker["artifact_contract"]["created_at"]
    safe_label = (label or "analyst_acceptance").replace(" ", "_")
    prefix = output_dir / f"unequal_roll_mvp_analyst_acceptance_tracker_{safe_label}_{timestamp}"

    paths = {
        "json": prefix.with_suffix(".json"),
        "markdown": prefix.with_suffix(".md"),
        "workbook": prefix.with_suffix(".xlsx"),
        "review_queue_csv": Path(f"{prefix}_review_queue.csv"),
        "summary_by_queue_csv": Path(f"{prefix}_summary_by_queue.csv"),
        "summary_by_county_csv": Path(f"{prefix}_summary_by_county.csv"),
        "summary_by_segment_csv": Path(f"{prefix}_summary_by_segment.csv"),
        "raw_input_index_csv": Path(f"{prefix}_raw_input_index.csv"),
    }

    paths["json"].write_text(json.dumps(tracker, indent=2, sort_keys=True))
    write_markdown(paths["markdown"], tracker)
    write_workbook(paths["workbook"], tracker)
    write_csv(paths["review_queue_csv"], tracker["review_queue_rows"], REVIEW_QUEUE_FIELDNAMES)
    write_csv(paths["summary_by_queue_csv"], tracker["summary"]["by_queue"], SUMMARY_QUEUE_FIELDNAMES)
    write_csv(paths["summary_by_county_csv"], tracker["summary"]["by_county"], SUMMARY_COUNTY_FIELDNAMES)
    write_csv(paths["summary_by_segment_csv"], tracker["summary"]["by_segment"], SUMMARY_SEGMENT_FIELDNAMES)
    write_csv(paths["raw_input_index_csv"], tracker["raw_input_index_rows"], RAW_INPUT_FIELDNAMES)
    return paths


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--packet-artifact",
        action="append",
        required=True,
        help="Packet JSON artifact path. Repeat or pass comma-separated paths.",
    )
    parser.add_argument("--output-dir", required=True, help="Directory for generated tracker artifacts.")
    parser.add_argument("--label", help="Optional label for generated files and packet rows.")
    parser.add_argument("--analyst-name", help="Optional default reviewer name to prefill.")
    parser.add_argument("--review-date", default=date.today().isoformat(), help="Optional default review date.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    packet_paths = _split_artifact_args(args.packet_artifact)
    tracker = build_tracker(
        packet_paths,
        label=args.label,
        analyst_name=args.analyst_name,
        review_date=args.review_date,
    )
    paths = write_outputs(tracker, Path(args.output_dir), args.label)
    print(json.dumps({key: str(path) for key, path in paths.items()}, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
