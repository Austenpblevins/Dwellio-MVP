#!/usr/bin/env python3
"""Build no-persist analyst pilot packets for governed simple reranking.

This report is deliberately downstream of validation. It does not rerun model
selection, change scoring, or persist results. Its job is to turn governed
`simple_value_tier_rerank` evidence into analyst-reviewable queue packets.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from statistics import median
from typing import Any

import psycopg
from psycopg.rows import dict_row

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - optional output format
    Workbook = None
    Alignment = Font = PatternFill = get_column_letter = None


VARIANT_KEY = "simple_value_tier_rerank"
ARCHITECTURE_NAME = "governed_similarity_baseline_with_simple_rerank"
MATERIAL_THRESHOLD = 1000.0
DEFAULT_DATABASE_URL = "postgresql://stage21_admin:stage21_admin@localhost:55442/stage21_dev"

SUBJECT_FACT_FIELDS = [
    "subject_address",
    "subject_appraised_value",
    "subject_living_area_sf",
    "subject_value_per_sf",
    "subject_year_built",
    "subject_land_sf",
    "subject_land_acres",
    "subject_bedrooms",
    "subject_full_baths",
    "subject_half_baths",
    "subject_stories",
    "subject_subdivision",
    "subject_neighborhood_code",
    "subject_quality_code",
    "subject_condition_code",
]

COMP_FACT_FIELDS = [
    "comp_tax_year",
    "comp_account_number",
    "comp_address",
    "comp_appraised_value",
    "comp_living_area_sf",
    "comp_value_per_sf",
    "comp_neighborhood_code",
    "comp_subdivision_name",
    "comp_land_sf",
    "comp_land_acres",
    "comp_bedrooms",
    "comp_full_baths",
    "comp_half_baths",
    "comp_stories",
    "comp_year_built",
    "comp_quality_code",
    "comp_condition_code",
]

SIGNOFF_FIELDNAMES = [
    "subject_account",
    "county_id",
    "neighborhood_code",
    "subject_address",
    "subject_appraised_value",
    "subject_value_per_sf",
    "subject_living_area_sf",
    "subject_year_built",
    "subject_land_sf",
    "subject_bedrooms",
    "subject_full_baths",
    "subject_half_baths",
    "subject_stories",
    "subject_quality_code",
    "subject_condition_code",
    "governed_taxpayer_savings",
    "smart_requested_roll_value",
    "rerank_requested_roll_value",
    "overlap_comp_count",
    "rerank_only_comp_count",
    "smart_only_comp_count",
    "final_decision",
    "analyst_decision",
    "comp_quality_issue",
    "neighborhood_or_subdivision_issue",
    "value_per_sf_issue",
    "removed_comp_was_better",
    "notes",
]

COMPARISON_FIELDS = [
    ("account", "Account"),
    ("county", "County"),
    ("neighborhood", "Neighborhood"),
    ("address", "Address"),
    ("tax_year", "Tax year"),
    ("appraised_value", "Appraised value"),
    ("value_per_sf", "Value/SF"),
    ("adjusted_value", "Adjusted Appraised Value"),
    ("adjusted_value_per_sf", "Adjusted Appraised Value/SF"),
    ("living_area_sf", "Living area SF"),
    ("year_built", "Year built"),
    ("effective_age", "Effective age"),
    ("land_sf", "Land SF"),
    ("bedrooms", "Beds"),
    ("full_baths", "Full baths"),
    ("half_baths", "Half baths"),
    ("stories", "Stories"),
    ("pool", "Pool"),
    ("quality", "Quality"),
    ("condition", "Condition"),
    ("subdivision", "Subdivision"),
    ("total_abs_adjustment", "Total adjustment burden"),
    ("line_item_count", "Line Item Count"),
]

COLUMN_KEY_ROWS = [
    {
        "field": "governed_taxpayer_savings",
        "plain_english": "Expected taxpayer-favorable value reduction retained by governed fallback versus similarity_top_100.",
        "analyst_action": "Use as benefit size; do not approve a case on savings alone.",
    },
    {
        "field": "membership",
        "plain_english": "Whether a comp is overlap, rerank_only, or smart_only.",
        "analyst_action": "Review rerank_only first, smart_only second, overlap third.",
    },
    {
        "field": "rerank_only",
        "plain_english": "Comp added by the simple reranker and not present in the smart baseline final comp set.",
        "analyst_action": "Primary evidence to inspect for credibility.",
    },
    {
        "field": "smart_only",
        "plain_english": "Comp removed by the simple reranker from the smart baseline final comp set.",
        "analyst_action": "Check whether removed comps were actually better than added comps.",
    },
    {
        "field": "overlap",
        "plain_english": "Comp retained by both smart baseline and simple reranker.",
        "analyst_action": "Use as context after changed comps are reviewed.",
    },
    {
        "field": "total_abs_adjustment",
        "plain_english": "Adjustment burden: summed absolute-dollar adjustments available in the source artifact.",
        "analyst_action": "Prefer lower burden; high burden needs analyst skepticism.",
    },
    {
        "field": "adjusted_value",
        "plain_english": "Comp adjusted appraised value from replay/final-value evidence when present.",
        "analyst_action": "Use only when populated by source evidence; blank means unavailable, not zero.",
    },
    {
        "field": "adjusted_value_per_sf",
        "plain_english": "Adjusted value divided by comp living area when adjusted value and living area are both available.",
        "analyst_action": "Do not infer this when the source adjusted value is blank.",
    },
    {
        "field": "median_appraised_value_per_sf",
        "plain_english": "Median unadjusted appraised value per SF across rerank final comps in the packet.",
        "analyst_action": "Use as context for the roll-based opinion; it is not a new model calculation.",
    },
    {
        "field": "adjusted_median_value_per_sf",
        "plain_english": "Rerank requested roll value divided by subject living area where both values are available.",
        "analyst_action": "Use as the packet's opinion value/SF summary, derived from existing rerank output.",
    },
    {
        "field": "opinion_of_value",
        "plain_english": "Roll-based unequal appraisal opinion of value from the governed rerank requested roll value.",
        "analyst_action": "Compare to current appraised value and confirm comps are defensible.",
    },
    {
        "field": "reduction_amount",
        "plain_english": "Current appraised value minus opinion of value where available, or rerank requested reduction from source output.",
        "analyst_action": "Use as expected value reduction; verify before external use.",
    },
    {
        "field": "reduction_percent",
        "plain_english": "Reduction amount divided by current appraised value when both values are available.",
        "analyst_action": "Use as sizing context, not a standalone approval reason.",
    },
    {
        "field": "model_backed",
        "plain_english": "Final value came from final_model_value rather than diagnostic/provisional paths.",
        "analyst_action": "First-pilot cases should be model-backed.",
    },
    {
        "field": "final_decision",
        "plain_english": (
            "Packet queue assignment: governed_rerank_ready, baseline_support_only, spot_check_only, "
            "analyst_review_only, hold_out, fallback_safety_blocked, or no_reduction_no_action."
        ),
        "analyst_action": "Use governed_rerank_ready for rerank wins and baseline_support_only for similarity_top_100 support.",
    },
    {
        "field": "governed_rerank_ready",
        "plain_english": "Simple rerank is model-backed, materially better than similarity_top_100, and passes safety checks.",
        "analyst_action": "Eligible for the rerank pilot review queue.",
    },
    {
        "field": "baseline_support_only",
        "plain_english": "similarity_top_100 supports a model-backed value reduction, but simple rerank did not add material benefit.",
        "analyst_action": "Review as baseline unequal-roll support; do not present as a rerank win.",
    },
    {
        "field": "fallback_safety_blocked",
        "plain_english": "Rerank and baseline support are not packet-ready because a safety blocker or missing evidence remains.",
        "analyst_action": "Do not use without further diagnosis.",
    },
    {
        "field": "no_reduction_no_action",
        "plain_english": "Neither governed rerank nor similarity_top_100 produced material model-backed value reduction.",
        "analyst_action": "No packet action.",
    },
]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build governed simple value-tier rerank analyst pilot packet."
    )
    parser.add_argument("--complete-comp-evidence-artifact", type=Path, required=True)
    parser.add_argument(
        "--governed-fallback-artifact",
        type=Path,
        action="append",
        default=[],
        help="Governed fallback artifact. Repeat for multiple cohorts.",
    )
    parser.add_argument(
        "--raw-artifact",
        type=Path,
        action="append",
        default=[],
        help="Raw true full-pool validation artifact. Repeat for multiple cohorts.",
    )
    parser.add_argument("--database-url", default=os.environ.get("DWELLIO_DATABASE_URL", DEFAULT_DATABASE_URL))
    parser.add_argument("--requested-tax-year", type=int, default=2026)
    parser.add_argument("--output-dir", type=Path, default=Path("/private/tmp"))
    parser.add_argument("--timestamp", default=None)
    return parser


def load_json(path: Path) -> dict[str, Any]:
    with path.open() as fh:
        return json.load(fh)


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    if fieldnames is None:
        fieldnames = []
        seen: set[str] = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    seen.add(key)
                    fieldnames.append(key)
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def as_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def as_int(value: Any, default: int = 0) -> int:
    if value in (None, ""):
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def optional_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in {"1", "true", "yes", "y"}
    return bool(value)


def decimal_to_builtin(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    return value


def money(value: float) -> str:
    return f"${value:,.2f}"


def safe_round(value: float | None, digits: int = 2) -> float | None:
    if value is None:
        return None
    return round(value, digits)


def median_or_none(values: list[float]) -> float | None:
    if not values:
        return None
    return float(median(values))


def case_key(row: dict[str, Any]) -> tuple[str, str, str]:
    return (
        str(row.get("county_id") or ""),
        str(row.get("subject_account") or ""),
        str(row.get("neighborhood_code") or ""),
    )


def split_ids(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(v) for v in value if str(v)]
    return [part.strip() for part in str(value).split(";") if part.strip()]


def normalize_membership(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if raw in {"rerank_only", "smart_only", "overlap"}:
        return raw
    return "unknown"


def summarize_cases(rows: list[dict[str, Any]]) -> dict[str, Any]:
    deltas = [
        as_float(row.get("packet_value_reduction_amount"), as_float(row.get("governed_taxpayer_delta_vs_similarity_top_100")))
        for row in rows
    ]
    return {
        "case_count": len(rows),
        "net_governed_taxpayer_savings": round(sum(deltas), 2),
        "model_backed_net_savings": round(
            sum(
                as_float(
                    row.get("packet_value_reduction_amount"),
                    as_float(row.get("governed_taxpayer_delta_vs_similarity_top_100")),
                )
                for row in rows
                if as_bool(row.get("model_backed")) or row.get("final_decision") == "baseline_support_only"
            ),
            2,
        ),
        "material_gain_count": sum(1 for delta in deltas if delta >= MATERIAL_THRESHOLD),
        "material_loss_count": sum(1 for delta in deltas if delta <= -MATERIAL_THRESHOLD),
        "median_savings": round(median(deltas), 2) if deltas else 0.0,
        "true_downgrade_count": sum(
            1
            for row in rows
            if row.get("final_decision") != "baseline_support_only"
            and as_bool(row.get("true_final_status_downgrade_raw"))
        ),
        "unsupported_transition_count": sum(
            1
            for row in rows
            if row.get("final_decision") != "baseline_support_only"
            and as_bool(row.get("true_transition_to_unsupported_raw"))
        ),
        "comp_collapse_count": sum(
            1
            for row in rows
            if row.get("final_decision") != "baseline_support_only"
            and as_bool(row.get("included_comp_collapse_raw"))
        ),
    }


def summarize_by(rows: list[dict[str, Any]], field: str) -> dict[str, dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row.get(field) or "")].append(row)
    return {key: summarize_cases(value) for key, value in sorted(grouped.items())}


def summarize_segments(rows: list[dict[str, Any]], limit: int | None = None) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[(str(row.get("county_id") or ""), str(row.get("neighborhood_code") or ""))].append(row)
    segment_rows = []
    for (county, neighborhood), values in grouped.items():
        summary = summarize_cases(values)
        summary.update({"county_id": county, "neighborhood_code": neighborhood})
        segment_rows.append(summary)
    segment_rows.sort(key=lambda row: (row["case_count"], row["net_governed_taxpayer_savings"]), reverse=True)
    if limit is not None:
        return segment_rows[:limit]
    return segment_rows


def build_raw_lookup(raw_artifact_paths: list[Path]) -> dict[tuple[str, str, str], dict[str, Any]]:
    lookup: dict[tuple[str, str, str], dict[str, Any]] = {}
    for path in raw_artifact_paths:
        if not path.exists():
            continue
        payload = load_json(path)
        for row in payload.get("subject_rows") or payload.get("variant_rows") or []:
            if row.get("variant_key") != VARIANT_KEY:
                continue
            lookup[(str(row.get("county_id")), str(row.get("subject_account")), str(row.get("neighborhood_code")))] = {
                "raw_source_artifact": str(path),
                "current_requested_roll_value": row.get("current_requested_roll_value"),
                "current_requested_reduction_amount": row.get("current_requested_reduction_amount"),
                "current_final_value_status": row.get("current_final_value_status"),
                "current_value_interpretation": row.get("current_value_interpretation"),
                "current_included_comp_count": row.get("current_included_comp_count"),
                "smart_requested_roll_value": row.get("smart_requested_roll_value"),
                "smart_requested_reduction_amount": row.get("smart_requested_reduction_amount"),
                "smart_final_value_status": row.get("smart_final_value_status"),
                "smart_value_interpretation": row.get("smart_value_interpretation"),
                "smart_included_comp_count_raw": row.get("smart_included_comp_count"),
                "rerank_requested_roll_value": row.get("rerank_requested_roll_value"),
                "rerank_requested_reduction_amount": row.get("rerank_requested_reduction_amount"),
                "rerank_final_value_status": row.get("rerank_final_value_status"),
                "rerank_value_interpretation_raw": row.get("rerank_value_interpretation"),
                "rerank_included_comp_count_raw": row.get("rerank_included_comp_count"),
                "rerank_vs_current_taxpayer_delta": row.get("rerank_vs_current_taxpayer_delta"),
                "rerank_vs_smart_taxpayer_delta": row.get("rerank_vs_smart_taxpayer_delta"),
            }
    return lookup


def load_fallback_blocked_rows(paths: list[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in paths:
        if not path.exists():
            continue
        payload = load_json(path)
        for row in payload.get("case_rows", []):
            if row.get("variant_key") != VARIANT_KEY:
                continue
            if row.get("governance_view") != "fallback_blocked":
                continue
            out = dict(row)
            out["source_governed_fallback_artifact"] = str(path)
            rows.append(out)
    return rows


def hydrate_subject_facts(
    rows: list[dict[str, Any]], database_url: str, requested_tax_year: int
) -> dict[str, dict[str, Any]]:
    parcel_ids = sorted({str(row.get("subject_parcel_id") or "").strip() for row in rows if row.get("subject_parcel_id")})
    if not parcel_ids:
        return {}
    with psycopg.connect(database_url, row_factory=dict_row, connect_timeout=5) as conn:
        conn.execute("SET SESSION CHARACTERISTICS AS TRANSACTION READ ONLY")
        with conn.cursor() as cur:
            cur.execute(
                """
                WITH requested AS (
                  SELECT unnest(%s::uuid[]) AS parcel_id
                ),
                subject_snapshot AS (
                  SELECT DISTINCT ON (pys.parcel_id)
                    pys.parcel_year_snapshot_id,
                    pys.parcel_id,
                    pys.county_id,
                    pys.tax_year,
                    pys.account_number
                  FROM parcel_year_snapshots AS pys
                  JOIN requested AS r
                    ON r.parcel_id = pys.parcel_id
                  WHERE pys.is_current = true
                    AND pys.tax_year = %s
                  ORDER BY pys.parcel_id, pys.tax_year DESC
                ),
                current_address AS (
                  SELECT DISTINCT ON (pa.parcel_id)
                    pa.parcel_id,
                    pa.situs_address,
                    pa.situs_city,
                    COALESCE(pa.situs_state, 'TX') AS situs_state,
                    pa.situs_zip
                  FROM parcel_addresses AS pa
                  JOIN subject_snapshot AS ss
                    ON ss.parcel_id = pa.parcel_id
                  WHERE pa.is_current = true
                  ORDER BY pa.parcel_id, pa.updated_at DESC, pa.created_at DESC, pa.parcel_address_id DESC
                )
                SELECT
                  ss.parcel_id::text AS subject_parcel_id,
                  ss.county_id AS subject_county_id,
                  ss.account_number AS subject_account,
                  concat_ws(
                    ', ',
                    COALESCE(ca.situs_address, p.situs_address),
                    COALESCE(ca.situs_city, p.situs_city),
                    concat_ws(' ', COALESCE(ca.situs_state, COALESCE(p.situs_state, 'TX')), COALESCE(ca.situs_zip, p.situs_zip))
                  ) AS subject_address,
                  COALESCE(pc.neighborhood_code, p.neighborhood_code) AS subject_neighborhood_code,
                  COALESCE(pc.subdivision_name, p.subdivision_name) AS subject_subdivision,
                  ass.appraised_value AS subject_appraised_value,
                  pi.living_area_sf AS subject_living_area_sf,
                  CASE
                    WHEN COALESCE(pi.living_area_sf, 0) > 0
                    THEN ROUND((ass.appraised_value::numeric / pi.living_area_sf::numeric), 2)
                    ELSE NULL
                  END AS subject_value_per_sf,
                  pi.year_built AS subject_year_built,
                  pl.land_sf AS subject_land_sf,
                  pl.land_acres AS subject_land_acres,
                  pi.bedrooms AS subject_bedrooms,
                  COALESCE(pi.full_baths, fbb.full_baths_derived) AS subject_full_baths,
                  COALESCE(pi.half_baths, fbb.half_baths_derived) AS subject_half_baths,
                  pi.stories AS subject_stories,
                  pi.quality_code AS subject_quality_code,
                  pi.condition_code AS subject_condition_code
                FROM subject_snapshot AS ss
                JOIN parcels AS p
                  ON p.parcel_id = ss.parcel_id
                LEFT JOIN current_address AS ca
                  ON ca.parcel_id = ss.parcel_id
                LEFT JOIN property_characteristics AS pc
                  ON pc.parcel_year_snapshot_id = ss.parcel_year_snapshot_id
                LEFT JOIN parcel_assessments AS ass
                  ON ass.parcel_id = ss.parcel_id
                 AND ass.tax_year = ss.tax_year
                LEFT JOIN parcel_improvements AS pi
                  ON pi.parcel_id = ss.parcel_id
                 AND pi.tax_year = ss.tax_year
                LEFT JOIN parcel_lands AS pl
                  ON pl.parcel_id = ss.parcel_id
                 AND pl.tax_year = ss.tax_year
                LEFT JOIN fort_bend_valuation_bathroom_features AS fbb
                  ON fbb.parcel_id = ss.parcel_id
                 AND fbb.tax_year = ss.tax_year
                ORDER BY ss.county_id, ss.account_number
                """,
                (parcel_ids, requested_tax_year),
            )
            return {
                str(row["subject_parcel_id"]): {key: decimal_to_builtin(value) for key, value in dict(row).items()}
                for row in cur.fetchall()
            }


def hydrate_comp_facts(
    comp_rows: list[dict[str, Any]], database_url: str, requested_tax_year: int
) -> dict[str, dict[str, Any]]:
    parcel_ids = sorted({str(row.get("comp_parcel_id") or "").strip() for row in comp_rows if row.get("comp_parcel_id")})
    if not parcel_ids:
        return {}
    with psycopg.connect(database_url, row_factory=dict_row, connect_timeout=5) as conn:
        conn.execute("SET SESSION CHARACTERISTICS AS TRANSACTION READ ONLY")
        with conn.cursor() as cur:
            cur.execute(
                """
                WITH requested AS (
                  SELECT unnest(%s::uuid[]) AS parcel_id
                ),
                comp_snapshot AS (
                  SELECT
                    pys.parcel_year_snapshot_id,
                    pys.parcel_id,
                    pys.county_id,
                    pys.tax_year,
                    pys.account_number
                  FROM parcel_year_snapshots AS pys
                  JOIN requested AS r
                    ON r.parcel_id = pys.parcel_id
                  WHERE pys.is_current = true
                    AND pys.tax_year = %s
                ),
                current_address AS (
                  SELECT DISTINCT ON (pa.parcel_id)
                    pa.parcel_id,
                    pa.situs_address,
                    pa.situs_city,
                    COALESCE(pa.situs_state, 'TX') AS situs_state,
                    pa.situs_zip
                  FROM parcel_addresses AS pa
                  JOIN comp_snapshot AS cs
                    ON cs.parcel_id = pa.parcel_id
                  WHERE pa.is_current = true
                  ORDER BY pa.parcel_id, pa.updated_at DESC, pa.created_at DESC, pa.parcel_address_id DESC
                )
                SELECT
                  cs.parcel_id::text AS comp_parcel_id,
                  cs.tax_year AS comp_tax_year,
                  cs.account_number AS comp_account_number,
                  concat_ws(
                    ', ',
                    COALESCE(ca.situs_address, p.situs_address),
                    COALESCE(ca.situs_city, p.situs_city),
                    concat_ws(' ', COALESCE(ca.situs_state, COALESCE(p.situs_state, 'TX')), COALESCE(ca.situs_zip, p.situs_zip))
                  ) AS comp_address,
                  ass.appraised_value AS comp_appraised_value,
                  pi.living_area_sf AS comp_living_area_sf,
                  CASE
                    WHEN COALESCE(pi.living_area_sf, 0) > 0
                    THEN ROUND((ass.appraised_value::numeric / pi.living_area_sf::numeric), 2)
                    ELSE NULL
                  END AS comp_value_per_sf,
                  COALESCE(pc.neighborhood_code, p.neighborhood_code) AS comp_neighborhood_code,
                  COALESCE(pc.subdivision_name, p.subdivision_name) AS comp_subdivision_name,
                  pl.land_sf AS comp_land_sf,
                  pl.land_acres AS comp_land_acres,
                  pi.bedrooms AS comp_bedrooms,
                  COALESCE(pi.full_baths, fbb.full_baths_derived) AS comp_full_baths,
                  COALESCE(pi.half_baths, fbb.half_baths_derived) AS comp_half_baths,
                  pi.stories AS comp_stories,
                  pi.year_built AS comp_year_built,
                  pi.quality_code AS comp_quality_code,
                  pi.condition_code AS comp_condition_code
                FROM comp_snapshot AS cs
                JOIN parcels AS p
                  ON p.parcel_id = cs.parcel_id
                LEFT JOIN current_address AS ca
                  ON ca.parcel_id = cs.parcel_id
                LEFT JOIN property_characteristics AS pc
                  ON pc.parcel_year_snapshot_id = cs.parcel_year_snapshot_id
                LEFT JOIN parcel_assessments AS ass
                  ON ass.parcel_id = cs.parcel_id
                 AND ass.tax_year = cs.tax_year
                LEFT JOIN parcel_improvements AS pi
                  ON pi.parcel_id = cs.parcel_id
                 AND pi.tax_year = cs.tax_year
                LEFT JOIN parcel_lands AS pl
                  ON pl.parcel_id = cs.parcel_id
                 AND pl.tax_year = cs.tax_year
                LEFT JOIN fort_bend_valuation_bathroom_features AS fbb
                  ON fbb.parcel_id = cs.parcel_id
                 AND fbb.tax_year = cs.tax_year
                ORDER BY cs.county_id, cs.account_number
                """,
                (parcel_ids, requested_tax_year),
            )
            return {
                str(row["comp_parcel_id"]): {key: decimal_to_builtin(value) for key, value in dict(row).items()}
                for row in cur.fetchall()
            }


def apply_subject_hydration(row: dict[str, Any], subject_facts: dict[str, dict[str, Any]]) -> None:
    facts = subject_facts.get(str(row.get("subject_parcel_id") or ""))
    if not facts:
        row["subject_hydration_status"] = "missing_source_row"
    else:
        row["subject_hydration_status"] = "hydrated"
        for field in SUBJECT_FACT_FIELDS:
            value = facts.get(field)
            if value not in (None, ""):
                row[field] = value
    if row.get("county_id") == "fort_bend":
        for field in ("subject_full_baths", "subject_half_baths"):
            if row.get(field) in (None, ""):
                row[field] = "source_unavailable"


def apply_comp_hydration(row: dict[str, Any], comp_facts: dict[str, dict[str, Any]]) -> None:
    row["original_reported_comp_tax_year"] = row.get("comp_tax_year")
    row["membership"] = normalize_membership(row.get("membership"))
    facts = comp_facts.get(str(row.get("comp_parcel_id") or ""))
    if not facts:
        row["comp_hydration_status"] = "missing_requested_tax_year_source_row"
    else:
        row["comp_hydration_status"] = f"hydrated_{facts.get('comp_tax_year')}"
        for field in COMP_FACT_FIELDS:
            value = facts.get(field)
            if value not in (None, ""):
                row[field] = value
    alias_pairs = {
        "candidate_address": "comp_address",
        "candidate_living_area_sf": "comp_living_area_sf",
        "candidate_land_sf": "comp_land_sf",
        "candidate_neighborhood_code": "comp_neighborhood_code",
        "candidate_subdivision_name": "comp_subdivision_name",
        "candidate_year_built": "comp_year_built",
        "candidate_bedrooms": "comp_bedrooms",
        "candidate_full_baths": "comp_full_baths",
        "candidate_half_baths": "comp_half_baths",
        "candidate_stories": "comp_stories",
        "candidate_quality_code": "comp_quality_code",
        "candidate_condition_code": "comp_condition_code",
    }
    for alias, source in alias_pairs.items():
        row[alias] = row.get(source)
    if row.get("county_id") == "fort_bend" or row.get("comp_county_id") == "fort_bend":
        for field in ("comp_full_baths", "comp_half_baths", "candidate_full_baths", "candidate_half_baths"):
            if row.get(field) in (None, ""):
                row[field] = "source_unavailable"


def baseline_value_reduction(row: dict[str, Any]) -> float:
    reduction = optional_float(row.get("smart_requested_reduction_amount"))
    if reduction is not None:
        return reduction
    appraised = optional_float(row.get("subject_appraised_value"))
    baseline_value = optional_float(row.get("smart_requested_roll_value"))
    if appraised is not None and baseline_value is not None:
        return appraised - baseline_value
    return 0.0


def baseline_is_model_backed(row: dict[str, Any]) -> bool:
    return str(row.get("smart_value_interpretation") or "").strip() == "final_model_value"


def baseline_status_is_supported(row: dict[str, Any]) -> bool:
    status = str(row.get("smart_final_value_status") or row.get("smart_replay_final_status") or "").strip()
    return status not in {"", "unsupported", "no_reduction"}


def comp_evidence_reasons(comp_rows: list[dict[str, Any]], requested_tax_year: int) -> list[str]:
    reasons: list[str] = []
    bad_tax_year_count = sum(1 for comp in comp_rows if str(comp.get("comp_tax_year") or "") != str(requested_tax_year))
    if bad_tax_year_count:
        reasons.append("wrong_tax_year_comp_rows")
    if any(comp.get("comp_hydration_status") == "missing_requested_tax_year_source_row" for comp in comp_rows):
        reasons.append("missing_requested_tax_year_comp_hydration")
    return reasons


def baseline_support_reasons(row: dict[str, Any], comp_rows: list[dict[str, Any]], requested_tax_year: int) -> list[str]:
    reasons: list[str] = []
    baseline_comp_rows = [
        comp for comp in comp_rows if normalize_membership(comp.get("membership")) in {"smart_only", "overlap"}
    ]
    if not baseline_is_model_backed(row):
        reasons.append("baseline_not_final_model_value")
    if not baseline_status_is_supported(row):
        reasons.append("baseline_final_status_not_supported")
    if baseline_value_reduction(row) < MATERIAL_THRESHOLD:
        reasons.append("baseline_reduction_below_material_threshold")
    if as_int(row.get("smart_included_comp_count") or row.get("smart_included_comp_count_raw")) <= 0:
        reasons.append("baseline_missing_included_comps")
    reasons.extend(comp_evidence_reasons(baseline_comp_rows, requested_tax_year))
    return reasons


def first_pilot_blocking_reasons(row: dict[str, Any], comp_rows: list[dict[str, Any]], requested_tax_year: int) -> list[str]:
    reasons: list[str] = []
    if row.get("governance_view") != "automated_safe":
        reasons.append("not_automated_safe_governance_view")
    if row.get("governance_classification") != "eligible_candidate":
        reasons.append("not_eligible_candidate")
    if not as_bool(row.get("model_backed")):
        reasons.append("not_model_backed")
    if as_bool(row.get("true_final_status_downgrade_raw")):
        reasons.append("true_final_status_downgrade")
    if as_bool(row.get("true_transition_to_unsupported_raw")):
        reasons.append("true_transition_to_unsupported")
    if as_bool(row.get("included_comp_collapse_raw")):
        reasons.append("included_comp_collapse")
    if str(row.get("rerank_replay_value_interpretation") or row.get("rerank_value_interpretation_raw") or "") != "final_model_value":
        reasons.append("rerank_not_final_model_value")
    if as_float(row.get("governed_taxpayer_delta_vs_similarity_top_100")) < MATERIAL_THRESHOLD:
        reasons.append("taxpayer_delta_below_material_threshold")
    reasons.extend(comp_evidence_reasons(comp_rows, requested_tax_year))
    return reasons


def pilot_decision(row: dict[str, Any], comp_rows: list[dict[str, Any]], requested_tax_year: int) -> tuple[str, list[str]]:
    blocking_reasons = first_pilot_blocking_reasons(row, comp_rows, requested_tax_year)
    baseline_reasons = baseline_support_reasons(row, comp_rows, requested_tax_year)
    has_baseline_support = not baseline_reasons
    if row.get("governance_view") == "fallback_blocked":
        if has_baseline_support:
            return "baseline_support_only", ["rerank_not_materially_better_baseline_support_retained"]
        if any(reason in set(blocking_reasons + baseline_reasons) for reason in ("wrong_tax_year_comp_rows", "missing_requested_tax_year_comp_hydration")):
            return "fallback_safety_blocked", blocking_reasons + baseline_reasons
        if row.get("governance_classification") in {"blocked_case", "insufficient_evidence"}:
            return "fallback_safety_blocked", blocking_reasons + baseline_reasons
        return "no_reduction_no_action", baseline_reasons or ["baseline_reduction_below_material_threshold"]
    if as_bool(row.get("true_final_status_downgrade_raw")):
        return "hold_out", blocking_reasons or ["true_final_status_downgrade"]
    if row.get("governance_view") == "analyst_assisted" or row.get("governance_classification") == "manual_review_required":
        return "analyst_review_only", blocking_reasons or ["manual_review_governance_view"]
    if blocking_reasons:
        return "analyst_review_only", blocking_reasons
    overlap = as_int(row.get("smart_vs_rerank_overlap_count") or row.get("overlap_comp_count_recovered"))
    if overlap < 8:
        return "analyst_review_only", ["weak_comp_set_overlap"]
    if overlap < 10:
        return "spot_check_only", ["moderate_comp_set_overlap_spot_check"]
    return "governed_rerank_ready", ["model_backed_stable_material_benefit_complete_evidence"]


def final_reason(decision: str, reasons: list[str]) -> str:
    if decision == "governed_rerank_ready":
        return (
            "Model-backed final-value result with complete 2026 smart/rerank comp evidence, material taxpayer "
            "savings, no downgrade, no unsupported transition, no comp collapse, and no pre-pilot QA flag."
        )
    if decision == "baseline_support_only":
        return (
            "similarity_top_100 produced model-backed baseline unequal-roll support, while simple rerank was tested "
            "but did not add material incremental savings."
        )
    if decision == "spot_check_only":
        return "Clean governed result held outside the first pilot queue for comp-overlap spot-check."
    if decision == "analyst_review_only":
        return f"Retained governed evidence requires analyst review before pilot use: {', '.join(reasons)}."
    if decision == "fallback_safety_blocked":
        return f"Fallback kept similarity_top_100, but packet support is safety-blocked: {', '.join(reasons)}."
    if decision == "no_reduction_no_action":
        return f"No material model-backed baseline or rerank value reduction: {', '.join(reasons)}."
    return f"Held out from first pilot: {', '.join(reasons)}."


def enrich_case_rows(
    case_rows: list[dict[str, Any]],
    raw_lookup: dict[tuple[str, str, str], dict[str, Any]],
) -> list[dict[str, Any]]:
    enriched = []
    for row in case_rows:
        if row.get("variant_key") != VARIANT_KEY:
            continue
        out = dict(row)
        out.update(raw_lookup.get(case_key(out), {}))
        out["smart_included_comp_ids"] = out.get("smart_full_included_comp_ids", "")
        out["rerank_included_comp_ids"] = out.get("rerank_full_included_comp_ids", "")
        out["rerank_only_comp_ids"] = out.get("added_comp_ids", "")
        out["smart_only_comp_ids"] = out.get("removed_comp_ids", "")
        out["analyst_review_required_for_pilot"] = True
        enriched.append(out)
    return enriched


def attach_subject_facts_to_comp(row: dict[str, Any], case_row: dict[str, Any]) -> None:
    for field in SUBJECT_FACT_FIELDS:
        row[field] = case_row.get(field)
    row["subject_hydration_status"] = case_row.get("subject_hydration_status")


def build_queue_rows(
    complete_payload: dict[str, Any],
    raw_lookup: dict[tuple[str, str, str], dict[str, Any]],
    database_url: str,
    requested_tax_year: int,
) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]], dict[str, Any]]:
    case_rows = enrich_case_rows(complete_payload.get("case_rows", []), raw_lookup)
    comp_rows = [dict(row) for row in complete_payload.get("comp_rows", [])]

    subject_facts = hydrate_subject_facts(case_rows, database_url, requested_tax_year)
    comp_facts = hydrate_comp_facts(comp_rows, database_url, requested_tax_year)

    for row in case_rows:
        row["subject_tax_year"] = requested_tax_year
        apply_subject_hydration(row, subject_facts)
    for row in comp_rows:
        apply_comp_hydration(row, comp_facts)

    comp_by_case: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    case_by_key = {case_key(row): row for row in case_rows}
    for comp in comp_rows:
        owner = case_by_key.get(case_key(comp))
        if not owner:
            continue
        attach_subject_facts_to_comp(comp, owner)
        comp["adjustment_burden_explanation"] = (
            "total_abs_adjustment is adjustment burden: the absolute-dollar adjustment amount summed across "
            "available adjustment line items for this comp. Lower values are generally easier to defend. "
            "The source artifacts expose total_abs_adjustment and line_item_count, not individual line items."
        )
        comp_by_case[case_key(comp)].append(comp)

    queues = {
        "governed_rerank_ready": [],
        "baseline_support_only": [],
        "spot_check_only": [],
        "analyst_review_only": [],
        "hold_out": [],
        "fallback_safety_blocked": [],
        "no_reduction_no_action": [],
    }
    for row in case_rows:
        decision, reasons = pilot_decision(row, comp_by_case.get(case_key(row), []), requested_tax_year)
        row["final_decision"] = decision
        row["final_decision_reasons"] = ";".join(reasons)
        row["reason_passed_or_excluded"] = final_reason(decision, reasons)
        row["baseline_value_reduction_amount"] = baseline_value_reduction(row)
        row["packet_value_reduction_amount"] = (
            baseline_value_reduction(row)
            if decision == "baseline_support_only"
            else as_float(row.get("governed_taxpayer_delta_vs_similarity_top_100"))
        )
        row["architecture_name"] = ARCHITECTURE_NAME
        row["analyst_review_required_for_pilot"] = decision != "governed_rerank_ready"
        queues[decision].append(row)

    recovery_summary = {
        "case_rows_loaded": len(case_rows),
        "comp_rows_loaded": len(comp_rows),
        "subject_facts_requested": len({str(row.get("subject_parcel_id") or "") for row in case_rows if row.get("subject_parcel_id")}),
        "subject_facts_hydrated": len(subject_facts),
        "comp_facts_requested": len({str(row.get("comp_parcel_id") or "") for row in comp_rows if row.get("comp_parcel_id")}),
        "comp_facts_hydrated": len(comp_facts),
        "comp_membership_counts": dict(Counter(normalize_membership(row.get("membership")) for row in comp_rows)),
        "comp_hydration_status_counts": dict(Counter(str(row.get("comp_hydration_status") or "") for row in comp_rows)),
    }
    return queues, comp_rows, recovery_summary


def filter_comp_rows(comp_rows: list[dict[str, Any]], case_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    keys = {case_key(row) for row in case_rows}
    decisions = {case_key(row): row.get("final_decision") for row in case_rows}
    return [dict(row, final_decision=decisions.get(case_key(row))) for row in comp_rows if case_key(row) in keys]


def build_signoff_rows(first_pilot_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for row in first_pilot_rows:
        rows.append(
            {
                "subject_account": row.get("subject_account"),
                "county_id": row.get("county_id"),
                "neighborhood_code": row.get("neighborhood_code"),
                "subject_address": row.get("subject_address"),
                "subject_appraised_value": row.get("subject_appraised_value"),
                "subject_value_per_sf": row.get("subject_value_per_sf"),
                "subject_living_area_sf": row.get("subject_living_area_sf"),
                "subject_year_built": row.get("subject_year_built"),
                "subject_land_sf": row.get("subject_land_sf"),
                "subject_bedrooms": row.get("subject_bedrooms"),
                "subject_full_baths": row.get("subject_full_baths"),
                "subject_half_baths": row.get("subject_half_baths"),
                "subject_stories": row.get("subject_stories"),
                "subject_quality_code": row.get("subject_quality_code"),
                "subject_condition_code": row.get("subject_condition_code"),
                "governed_taxpayer_savings": row.get("packet_value_reduction_amount") or row.get("governed_taxpayer_delta_vs_similarity_top_100"),
                "smart_requested_roll_value": row.get("smart_requested_roll_value"),
                "rerank_requested_roll_value": row.get("rerank_requested_roll_value"),
                "overlap_comp_count": row.get("smart_vs_rerank_overlap_count"),
                "rerank_only_comp_count": len(split_ids(row.get("rerank_only_comp_ids"))),
                "smart_only_comp_count": len(split_ids(row.get("smart_only_comp_ids"))),
                "final_decision": row.get("final_decision"),
                "analyst_decision": "",
                "comp_quality_issue": "",
                "neighborhood_or_subdivision_issue": "",
                "value_per_sf_issue": "",
                "removed_comp_was_better": "",
                "notes": "",
            }
        )
    return rows


def adjusted_value_per_sf(comp: dict[str, Any]) -> float | None:
    adjusted = optional_float(comp.get("adjusted_value"))
    living_area = optional_float(comp.get("comp_living_area_sf"))
    if adjusted is None or not living_area or living_area <= 0:
        return None
    return round(adjusted / living_area, 2)


def comp_display_value(comp: dict[str, Any], field: str) -> Any:
    if field == "adjusted_value_per_sf":
        return adjusted_value_per_sf(comp)
    mapping = {
        "account": "comp_account_number",
        "county": "comp_county_id",
        "neighborhood": "comp_neighborhood_code",
        "address": "comp_address",
        "tax_year": "comp_tax_year",
        "appraised_value": "comp_appraised_value",
        "value_per_sf": "comp_value_per_sf",
        "adjusted_value": "adjusted_value",
        "living_area_sf": "comp_living_area_sf",
        "year_built": "comp_year_built",
        "effective_age": "comp_effective_age",
        "land_sf": "comp_land_sf",
        "bedrooms": "comp_bedrooms",
        "full_baths": "comp_full_baths",
        "half_baths": "comp_half_baths",
        "stories": "comp_stories",
        "pool": "comp_pool_flag",
        "quality": "comp_quality_code",
        "condition": "comp_condition_code",
        "subdivision": "comp_subdivision_name",
        "total_abs_adjustment": "total_abs_adjustment",
        "line_item_count": "line_item_count",
    }
    return comp.get(mapping[field])


def subject_display_value(row: dict[str, Any], field: str) -> Any:
    mapping = {
        "account": "subject_account",
        "county": "county_id",
        "neighborhood": "neighborhood_code",
        "address": "subject_address",
        "tax_year": "subject_tax_year",
        "appraised_value": "subject_appraised_value",
        "value_per_sf": "subject_value_per_sf",
        "adjusted_value": None,
        "adjusted_value_per_sf": None,
        "living_area_sf": "subject_living_area_sf",
        "year_built": "subject_year_built",
        "effective_age": None,
        "land_sf": "subject_land_sf",
        "bedrooms": "subject_bedrooms",
        "full_baths": "subject_full_baths",
        "half_baths": "subject_half_baths",
        "stories": "subject_stories",
        "pool": None,
        "quality": "subject_quality_code",
        "condition": "subject_condition_code",
        "subdivision": "subject_subdivision",
        "total_abs_adjustment": None,
        "line_item_count": None,
    }
    source = mapping[field]
    return row.get(source) if source else ""


def build_changed_comp_rows(
    first_pilot_rows: list[dict[str, Any]], first_comp_rows: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    by_case = {case_key(row): row for row in first_pilot_rows}
    rows = []
    for comp in first_comp_rows:
        membership = normalize_membership(comp.get("membership"))
        if membership == "overlap":
            continue
        case = by_case.get(case_key(comp), {})
        rows.append(
            {
                "subject_account": comp.get("subject_account"),
                "county_id": comp.get("county_id"),
                "neighborhood_code": comp.get("neighborhood_code"),
                "subject_address": comp.get("subject_address"),
                "governed_taxpayer_savings": case.get("packet_value_reduction_amount") or case.get("governed_taxpayer_delta_vs_similarity_top_100"),
                "overlap_comp_count": case.get("smart_vs_rerank_overlap_count"),
                "membership": membership,
                "comp_account_number": comp.get("comp_account_number"),
                "comp_address": comp.get("comp_address"),
                "comp_tax_year": comp.get("comp_tax_year"),
                "comp_appraised_value": comp.get("comp_appraised_value"),
                "comp_value_per_sf": comp.get("comp_value_per_sf"),
                "adjusted_value": comp.get("adjusted_value"),
                "adjusted_value_per_sf": adjusted_value_per_sf(comp),
                "comp_living_area_sf": comp.get("comp_living_area_sf"),
                "comp_year_built": comp.get("comp_year_built"),
                "comp_effective_age": comp.get("comp_effective_age"),
                "comp_land_sf": comp.get("comp_land_sf"),
                "comp_bedrooms": comp.get("comp_bedrooms"),
                "comp_full_baths": comp.get("comp_full_baths"),
                "comp_half_baths": comp.get("comp_half_baths"),
                "comp_stories": comp.get("comp_stories"),
                "comp_pool_flag": comp.get("comp_pool_flag"),
                "comp_quality_code": comp.get("comp_quality_code"),
                "comp_condition_code": comp.get("comp_condition_code"),
                "comp_subdivision_name": comp.get("comp_subdivision_name"),
                "total_abs_adjustment": comp.get("total_abs_adjustment"),
                "line_item_count": comp.get("line_item_count"),
                "analyst_review_priority": "1_review_added_rerank_comp" if membership == "rerank_only" else "2_review_removed_smart_comp",
                "analyst_notes": "",
            }
        )
    rows.sort(
        key=lambda row: (
            row["subject_account"],
            0 if row["membership"] == "rerank_only" else 1,
            str(row.get("comp_account_number") or ""),
        )
    )
    return rows


def build_comparison_grid_rows(
    first_pilot_rows: list[dict[str, Any]], first_comp_rows: list[dict[str, Any]], max_comps: int = 20
) -> list[dict[str, Any]]:
    comps_by_case: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for comp in first_comp_rows:
        comps_by_case[case_key(comp)].append(comp)
    rows: list[dict[str, Any]] = []
    membership_order = {"rerank_only": 0, "overlap": 1, "smart_only": 2}
    for case in first_pilot_rows:
        comps = sorted(
            comps_by_case.get(case_key(case), []),
            key=lambda comp: (
                membership_order.get(normalize_membership(comp.get("membership")), 9),
                str(comp.get("comp_account_number") or ""),
            ),
        )[:max_comps]
        for field, label in COMPARISON_FIELDS:
            out: dict[str, Any] = {
                "subject_account": case.get("subject_account"),
                "county_id": case.get("county_id"),
                "neighborhood_code": case.get("neighborhood_code"),
                "governed_taxpayer_savings": case.get("packet_value_reduction_amount") or case.get("governed_taxpayer_delta_vs_similarity_top_100"),
                "row_label": label,
                "SUBJECT": subject_display_value(case, field),
            }
            for index, comp in enumerate(comps, start=1):
                prefix = f"{normalize_membership(comp.get('membership')).upper()} COMP {index}"
                out[prefix] = comp_display_value(comp, field)
            rows.append(out)
        rows.append(
            {
                "subject_account": case.get("subject_account"),
                "county_id": case.get("county_id"),
                "neighborhood_code": case.get("neighborhood_code"),
                "governed_taxpayer_savings": case.get("packet_value_reduction_amount") or case.get("governed_taxpayer_delta_vs_similarity_top_100"),
                "row_label": "",
                "SUBJECT": "",
            }
        )
    return rows


def build_opinion_of_value_rows(
    first_pilot_rows: list[dict[str, Any]], first_comp_rows: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    comps_by_case: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for comp in first_comp_rows:
        comps_by_case[case_key(comp)].append(comp)

    rows: list[dict[str, Any]] = []
    for case in first_pilot_rows:
        allowed_memberships = (
            {"smart_only", "overlap"}
            if case.get("final_decision") == "baseline_support_only"
            else {"rerank_only", "overlap"}
        )
        comps = sorted(
            [
                comp
                for comp in comps_by_case.get(case_key(case), [])
                if normalize_membership(comp.get("membership")) in allowed_memberships
            ],
            key=lambda comp: (
                0 if normalize_membership(comp.get("membership")) == "rerank_only" else 1,
                str(comp.get("comp_account_number") or ""),
            ),
        )
        appraised_vpsf_values = [
            value
            for value in (optional_float(comp.get("comp_value_per_sf")) for comp in comps)
            if value is not None
        ]
        adjusted_vpsf_values = [
            value
            for value in (adjusted_value_per_sf(comp) for comp in comps)
            if value is not None
        ]
        median_appraised_vpsf = safe_round(median_or_none(appraised_vpsf_values))
        adjusted_median_vpsf = safe_round(median_or_none(adjusted_vpsf_values))
        subject_living_area = optional_float(case.get("subject_living_area_sf"))
        opinion_source = "similarity_top_100_baseline" if case.get("final_decision") == "baseline_support_only" else "governed_simple_rerank"
        opinion_value = optional_float(
            case.get("smart_requested_roll_value")
            if case.get("final_decision") == "baseline_support_only"
            else case.get("rerank_requested_roll_value")
        )
        if adjusted_median_vpsf is None and opinion_value is not None and subject_living_area:
            adjusted_median_vpsf = round(opinion_value / subject_living_area, 2)
        subject_appraised = optional_float(case.get("subject_appraised_value"))
        reduction_amount = optional_float(
            case.get("smart_requested_reduction_amount")
            if case.get("final_decision") == "baseline_support_only"
            else case.get("rerank_requested_reduction_amount")
        )
        if reduction_amount is None and subject_appraised is not None and opinion_value is not None:
            reduction_amount = subject_appraised - opinion_value
        reduction_percent = None
        if reduction_amount is not None and subject_appraised and subject_appraised > 0:
            reduction_percent = round(reduction_amount / subject_appraised, 4)

        for index, comp in enumerate(comps, start=1):
            rows.append(
                {
                    "subject_account": case.get("subject_account"),
                    "subject_address": case.get("subject_address"),
                    "county_id": case.get("county_id"),
                    "neighborhood_code": case.get("neighborhood_code"),
                    "subject_living_area_sf": case.get("subject_living_area_sf"),
                    "subject_current_appraised_value": case.get("subject_appraised_value"),
                    "subject_value_per_sf": case.get("subject_value_per_sf"),
                    "opinion_of_value": opinion_value,
                    "opinion_source": opinion_source,
                    "reduction_amount": reduction_amount,
                    "reduction_percent": reduction_percent,
                    "median_appraised_value_per_sf": median_appraised_vpsf,
                    "adjusted_median_value_per_sf": adjusted_median_vpsf,
                    "comp_number": index,
                    "membership": normalize_membership(comp.get("membership")),
                    "comp_account_number": comp.get("comp_account_number"),
                    "comp_address": comp.get("comp_address"),
                    "comp_appraised_value": comp.get("comp_appraised_value"),
                    "comp_appraised_value_per_sf": comp.get("comp_value_per_sf"),
                    "comp_adjusted_value": comp.get("adjusted_value"),
                    "comp_adjusted_value_per_sf": adjusted_value_per_sf(comp),
                    "comp_living_area_sf": comp.get("comp_living_area_sf"),
                    "total_abs_adjustment": comp.get("total_abs_adjustment"),
                    "line_item_count": comp.get("line_item_count"),
                    "adjusted_value_source_note": (
                        "source_artifact_adjusted_value"
                        if comp.get("adjusted_value") not in (None, "")
                        else "unavailable_in_source_artifact"
                    ),
                }
            )
    return rows


def build_pilot_summary_rows(summary: dict[str, Any]) -> list[dict[str, Any]]:
    rows = [
        {"metric": "Architecture", "value": summary["architecture_name"]},
        {"metric": "Governed-rerank-ready count", "value": summary["governed_rerank_ready"]["case_count"]},
        {"metric": "Baseline-support-only count", "value": summary["baseline_support_only"]["case_count"]},
        {"metric": "Expected model-backed rerank value reduction", "value": summary["governed_rerank_ready"]["model_backed_net_savings"]},
        {"metric": "Expected model-backed baseline value reduction", "value": summary["baseline_support_only"]["model_backed_net_savings"]},
        {"metric": "First-pilot-ready count (compatibility alias)", "value": summary["first_pilot_ready"]["case_count"]},
        {"metric": "Material gains", "value": summary["first_pilot_ready"]["material_gain_count"]},
        {"metric": "Material losses", "value": summary["first_pilot_ready"]["material_loss_count"]},
        {"metric": "Downgrades", "value": summary["first_pilot_ready"]["true_downgrade_count"]},
        {"metric": "Unsupported transitions", "value": summary["first_pilot_ready"]["unsupported_transition_count"]},
        {"metric": "Comp collapses", "value": summary["first_pilot_ready"]["comp_collapse_count"]},
        {"metric": "First-pilot comp rows outside requested tax year", "value": summary["comp_hydration_summary"]["first_pilot_non_requested_tax_year_comp_rows"]},
        {"metric": "First-pilot subject hydration", "value": f"{summary['subject_hydration_summary']['first_pilot_hydrated_count']} / {summary['subject_hydration_summary']['first_pilot_case_count']}"},
    ]
    for county, stats in summary["first_pilot_county_summary"].items():
        rows.append({"metric": f"{county} first-pilot count", "value": stats["case_count"]})
        rows.append({"metric": f"{county} expected value reduction", "value": stats["model_backed_net_savings"]})
    for item in summary["pilot_acceptance_framework"]:
        rows.append({"metric": "Acceptance criterion", "value": item})
    return rows


def maybe_write_workbook(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> bool:
    if Workbook is None:
        return False
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet_name, rows in sheets.items():
        sheet = workbook.create_sheet(sheet_name[:31])
        if not rows:
            sheet.append(["No rows"])
            continue
        fieldnames: list[str] = []
        seen: set[str] = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    seen.add(key)
                    fieldnames.append(key)
        sheet.append(fieldnames)
        for row in rows:
            sheet.append([excel_cell_value(row.get(field)) for field in fieldnames])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_index, field in enumerate(fieldnames, start=1):
            width = min(max(len(field) + 2, 12), 42)
            for value in [row.get(field) for row in rows[:100]]:
                width = min(max(width, min(len(str(value or "")) + 2, 42)), 42)
            sheet.column_dimensions[get_column_letter(column_index)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(path)
    return True


def excel_cell_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (list, dict)):
        return json.dumps(value, sort_keys=True)
    return value


def build_fallback_blocked_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return summarize_cases([])
    return {
        **summarize_cases(rows),
        "fallback_count": sum(1 for row in rows if as_bool(row.get("fallback_to_similarity_top_100"))),
        "fallback_prevented_harm_count": sum(1 for row in rows if as_bool(row.get("fallback_prevented_harm"))),
        "fallback_removed_apparent_savings_count": sum(1 for row in rows if as_bool(row.get("fallback_removed_apparent_savings"))),
    }


def build_summary(
    queues: dict[str, list[dict[str, Any]]],
    comp_first_pilot: list[dict[str, Any]],
    comp_baseline_support: list[dict[str, Any]],
    comp_review_packet: list[dict[str, Any]],
    fallback_blocked: list[dict[str, Any]],
    recovery_summary: dict[str, Any],
    requested_tax_year: int,
) -> dict[str, Any]:
    first = queues["governed_rerank_ready"]
    baseline = queues["baseline_support_only"]
    all_rows = (
        queues["governed_rerank_ready"]
        + queues["baseline_support_only"]
        + queues["spot_check_only"]
        + queues["analyst_review_only"]
        + queues["hold_out"]
        + queues["fallback_safety_blocked"]
        + queues["no_reduction_no_action"]
    )
    non_requested_tax_year_rows = [
        row for row in comp_first_pilot if str(row.get("comp_tax_year") or "") != str(requested_tax_year)
    ]
    review_non_requested_tax_year_rows = [
        row for row in comp_review_packet if str(row.get("comp_tax_year") or "") != str(requested_tax_year)
    ]
    return {
        "architecture_name": ARCHITECTURE_NAME,
        "governed_rerank_ready": summarize_cases(first),
        "first_pilot_ready": summarize_cases(first),
        "baseline_support_only": summarize_cases(baseline),
        "spot_check_only": summarize_cases(queues["spot_check_only"]),
        "analyst_review_only": summarize_cases(queues["analyst_review_only"]),
        "hold_out": summarize_cases(queues["hold_out"]),
        "fallback_safety_blocked": summarize_cases(queues["fallback_safety_blocked"]),
        "no_reduction_no_action": summarize_cases(queues["no_reduction_no_action"]),
        "fallback_blocked": build_fallback_blocked_summary(fallback_blocked),
        "decision_counts": dict(Counter(row.get("final_decision") for row in all_rows)),
        "first_pilot_county_summary": summarize_by(first, "county_id"),
        "baseline_support_county_summary": summarize_by(baseline, "county_id"),
        "first_pilot_segment_summary": summarize_segments(first),
        "baseline_support_segment_summary": summarize_segments(baseline),
        "top_first_pilot_segments": summarize_segments(first, limit=15),
        "subject_hydration_summary": {
            "first_pilot_case_count": len(first),
            "first_pilot_hydrated_count": sum(1 for row in first if row.get("subject_hydration_status") == "hydrated"),
            "all_retained_case_count": len(all_rows),
            "all_retained_hydrated_count": sum(1 for row in all_rows if row.get("subject_hydration_status") == "hydrated"),
            "missing_subject_fields_first_pilot": {
                field: sum(1 for row in first if row.get(field) in (None, ""))
                for field in SUBJECT_FACT_FIELDS
            },
        },
        "comp_hydration_summary": {
            **recovery_summary,
            "first_pilot_comp_rows": len(comp_first_pilot),
            "baseline_support_comp_rows": len(comp_baseline_support),
            "review_packet_comp_rows": len(comp_review_packet),
            "first_pilot_non_requested_tax_year_comp_rows": len(non_requested_tax_year_rows),
            "review_packet_non_requested_tax_year_comp_rows": len(review_non_requested_tax_year_rows),
            "first_pilot_source_unavailable_comp_full_baths": sum(
                1 for row in comp_first_pilot if row.get("comp_full_baths") in (None, "", "source_unavailable")
            ),
        },
        "pilot_acceptance_framework": [
            "Target analyst approval rate for first-pilot-ready queue: at least 85%.",
            "Maximum allowed material analyst-rejected first-pilot cases before pausing expansion: 3.",
            "Maximum allowed Fort Bend analyst-rejected first-pilot cases before county-specific pause: 2.",
            "Any discovered unsupported transition, comp collapse, downgrade, or wrong-tax-year comp issue in the first-pilot queue pauses automation expansion.",
            "Spot-check, analyst-review, hold-out, and fallback-blocked queues are appendices only; they are not first-pilot output.",
            "Pilot decisions remain no-persist and analyst-reviewed; no production defaults change.",
        ],
        "analyst_review_instructions": {
            "membership_review_priority": ["rerank_only", "smart_only", "overlap"],
            "subject_level_issue_columns": "Mark yes/no on the subject row when an issue affects the case.",
            "multiple_comp_issue_instruction": "If multiple comps have issues, list comp account numbers and reasons in notes.",
            "adjustment_burden_explanation": (
                "total_abs_adjustment is total adjustment burden. Lower values are generally easier to defend. "
                "Line-item adjustment detail is not available in the source artifact."
            ),
        },
    }


def render_markdown(path: Path, payload: dict[str, Any]) -> None:
    summary = payload["summary"]
    first = payload["first_pilot_case_rows"]
    lines = [
        "# Governed Simple Value-Tier Rerank: Analyst Pilot Packet",
        "",
        "This is no-persist MVP pilot tooling, not production rollout. The architecture is "
        f"`{ARCHITECTURE_NAME}`: run `similarity_top_100` baseline support, attempt "
        "`simple_value_tier_rerank`, retain the rerank only when it materially and safely improves the baseline, "
        "otherwise fall back to baseline support when the baseline itself is model-backed.",
        "",
        "## Pilot Scope",
        "",
        f"- Governed-rerank-ready cases: {summary['governed_rerank_ready']['case_count']}",
        f"- Baseline-support-only cases: {summary['baseline_support_only']['case_count']}",
        f"- Expected model-backed rerank taxpayer savings: {money(summary['governed_rerank_ready']['model_backed_net_savings'])}",
        f"- Expected model-backed baseline value reduction: {money(summary['baseline_support_only']['model_backed_net_savings'])}",
        f"- Rerank material gains/losses: {summary['governed_rerank_ready']['material_gain_count']} / {summary['governed_rerank_ready']['material_loss_count']}",
        f"- Rerank downgrades / unsupported transitions / comp collapses: {summary['governed_rerank_ready']['true_downgrade_count']} / {summary['governed_rerank_ready']['unsupported_transition_count']} / {summary['governed_rerank_ready']['comp_collapse_count']}",
        "",
        "## Separated Queues",
        "",
        f"- Spot-check only: {summary['spot_check_only']['case_count']}",
        f"- Analyst-review only: {summary['analyst_review_only']['case_count']}",
        f"- Hold-out: {summary['hold_out']['case_count']}",
        f"- Fallback safety-blocked: {summary['fallback_safety_blocked']['case_count']}",
        f"- No reduction/no action: {summary['no_reduction_no_action']['case_count']}",
        f"- Raw governed fallback-blocked appendix rows: {summary['fallback_blocked']['case_count']}",
        "",
        "## County Scope",
        "",
        "| County | Cases | Expected savings | Median savings |",
        "|---|---:|---:|---:|",
    ]
    for county, stats in summary["first_pilot_county_summary"].items():
        lines.append(
            f"| {county} | {stats['case_count']} | {money(stats['net_governed_taxpayer_savings'])} | {money(stats['median_savings'])} |"
        )
    lines.extend(
        [
            "",
            "## Top First-Pilot Segments",
            "",
            "| County | Neighborhood | Cases | Expected savings |",
            "|---|---|---:|---:|",
        ]
    )
    for row in summary["top_first_pilot_segments"]:
        lines.append(
            f"| {row['county_id']} | `{row['neighborhood_code']}` | {row['case_count']} | {money(row['net_governed_taxpayer_savings'])} |"
        )
    lines.extend(
        [
            "",
            "## Analyst Checklist",
            "",
            "- Review comp membership in this order: `rerank_only`, `smart_only`, `overlap`.",
            "- Compare subject facts to rerank-only comps before approving.",
            "- Reject if a removed smart-only comp is plainly better than the added rerank-only comp set.",
            "- Check value per SF, living area, land size, year built, subdivision, address/location, and quality/condition where available.",
            "- Mark subject-level issue columns yes/no. If multiple comps have issues, list comp account numbers and reasons in `notes`.",
            "- Any discovered unsupported transition, comp collapse, downgrade, or wrong-tax-year comp issue pauses expansion.",
            "",
            "## Packet QA",
            "",
            f"- First-pilot comp rows outside requested tax year: {summary['comp_hydration_summary']['first_pilot_non_requested_tax_year_comp_rows']}",
            f"- First-pilot subject rows hydrated: {summary['subject_hydration_summary']['first_pilot_hydrated_count']} / {summary['subject_hydration_summary']['first_pilot_case_count']}",
            f"- Comp hydration status counts: `{summary['comp_hydration_summary']['comp_hydration_status_counts']}`",
            "",
            "## Top First-Pilot Cases",
            "",
            "| County | Account | Neighborhood | Savings | Smart status | Rerank status | Overlap |",
            "|---|---|---|---:|---|---|---:|",
        ]
    )
    for row in sorted(first, key=lambda r: as_float(r.get("packet_value_reduction_amount")), reverse=True)[:25]:
        lines.append(
            f"| {row.get('county_id')} | `{row.get('subject_account')}` | `{row.get('neighborhood_code')}` | "
            f"{money(as_float(row.get('packet_value_reduction_amount')))} | "
            f"{row.get('smart_replay_final_status') or row.get('smart_final_value_status')} | "
            f"{row.get('rerank_replay_final_status') or row.get('rerank_final_value_status')} | "
            f"{row.get('smart_vs_rerank_overlap_count')} |"
        )
    lines.extend(
        [
            "",
            "## Guardrails",
            "",
            "- No DB writes",
            "- No migrations",
            "- No production scoring, adjustment, median, governance, or final-value changes",
            "- No runtime default changes",
            "- No new penalties or reranking features",
            "- Reranking remains no-persist and experiment-only",
            "- Bounded proxy was not used for conclusions",
        ]
    )
    path.write_text("\n".join(lines) + "\n")


def output_paths(output_dir: Path, timestamp: str) -> dict[str, Path]:
    prefix = output_dir / f"unequal_roll_governed_simple_rerank_mvp_pilot_packet_{timestamp}"
    return {
        "json": prefix.with_suffix(".json"),
        "csv": prefix.with_suffix(".csv"),
        "markdown": prefix.with_suffix(".md"),
        "workbook": prefix.with_suffix(".xlsx"),
        "column_key": Path(f"{prefix}_column_key.csv"),
        "pilot_summary": Path(f"{prefix}_pilot_summary.csv"),
        "opinion_of_value": Path(f"{prefix}_opinion_of_value.csv"),
        "governed_rerank_ready": Path(f"{prefix}_governed_rerank_ready.csv"),
        "first_pilot": Path(f"{prefix}_first_pilot_ready.csv"),
        "first_pilot_comp_details": Path(f"{prefix}_first_pilot_comp_details.csv"),
        "baseline_support": Path(f"{prefix}_baseline_support_only.csv"),
        "baseline_support_comp_details": Path(f"{prefix}_baseline_support_comp_details.csv"),
        "signoff_tracker": Path(f"{prefix}_analyst_signoff_tracker.csv"),
        "comparison_grid": Path(f"{prefix}_subject_comp_comparison_grid.csv"),
        "changed_comps_review": Path(f"{prefix}_changed_comps_review.csv"),
        "spot_check": Path(f"{prefix}_spot_check_appendix.csv"),
        "analyst_review": Path(f"{prefix}_analyst_review_only.csv"),
        "hold_out": Path(f"{prefix}_hold_out.csv"),
        "fallback_safety_blocked": Path(f"{prefix}_fallback_safety_blocked.csv"),
        "no_reduction_no_action": Path(f"{prefix}_no_reduction_no_action.csv"),
        "fallback_blocked": Path(f"{prefix}_fallback_blocked.csv"),
        "excluded_comp_details": Path(f"{prefix}_excluded_queue_comp_details.csv"),
    }


def build_packet(
    complete_comp_evidence_artifact: Path,
    governed_fallback_artifacts: list[Path],
    raw_artifacts: list[Path],
    database_url: str,
    requested_tax_year: int,
    output_dir: Path,
    timestamp: str | None = None,
) -> tuple[dict[str, Any], dict[str, Path]]:
    complete_payload = load_json(complete_comp_evidence_artifact)
    raw_lookup = build_raw_lookup(raw_artifacts)
    queues, comp_rows, recovery_summary = build_queue_rows(
        complete_payload=complete_payload,
        raw_lookup=raw_lookup,
        database_url=database_url,
        requested_tax_year=requested_tax_year,
    )
    fallback_blocked = load_fallback_blocked_rows(governed_fallback_artifacts)
    first_comp = filter_comp_rows(comp_rows, queues["governed_rerank_ready"])
    baseline_comp = filter_comp_rows(comp_rows, queues["baseline_support_only"])
    review_packet_rows = queues["governed_rerank_ready"] + queues["baseline_support_only"]
    review_packet_comp = filter_comp_rows(comp_rows, review_packet_rows)
    excluded_comp = filter_comp_rows(
        comp_rows,
        queues["spot_check_only"]
        + queues["analyst_review_only"]
        + queues["hold_out"]
        + queues["fallback_safety_blocked"]
        + queues["no_reduction_no_action"],
    )
    summary = build_summary(
        queues=queues,
        comp_first_pilot=first_comp,
        comp_baseline_support=baseline_comp,
        comp_review_packet=review_packet_comp,
        fallback_blocked=fallback_blocked,
        recovery_summary=recovery_summary,
        requested_tax_year=requested_tax_year,
    )
    all_case_rows = (
        queues["governed_rerank_ready"]
        + queues["baseline_support_only"]
        + queues["spot_check_only"]
        + queues["analyst_review_only"]
        + queues["hold_out"]
        + queues["fallback_safety_blocked"]
        + queues["no_reduction_no_action"]
    )
    signoff_rows = build_signoff_rows(review_packet_rows)
    changed_comp_rows = build_changed_comp_rows(queues["governed_rerank_ready"], first_comp)
    comparison_grid_rows = build_comparison_grid_rows(review_packet_rows, review_packet_comp)
    opinion_rows = build_opinion_of_value_rows(review_packet_rows, review_packet_comp)
    pilot_summary_rows = build_pilot_summary_rows(summary)
    payload = {
        "artifact_contract": {
            "artifact_type": "governed_simple_value_tier_rerank_mvp_pilot_packet",
            "created_at": timestamp or datetime.now().strftime("%Y%m%dT%H%M%S"),
            "primary_variant": VARIANT_KEY,
            "architecture_name": ARCHITECTURE_NAME,
            "scope": "no_persist_mvp_pilot_tooling",
            "requested_tax_year": requested_tax_year,
            "bounded_proxy_used_for_conclusions": False,
        },
        "source_artifacts": {
            "complete_comp_evidence_artifact": str(complete_comp_evidence_artifact),
            "governed_fallback_artifacts": [str(path) for path in governed_fallback_artifacts],
            "raw_artifacts": [str(path) for path in raw_artifacts],
        },
        "guardrails": {
            "db_writes": False,
            "migrations": False,
            "runtime_defaults_changed": False,
            "production_scoring_adjustment_median_governance_final_value_changed": False,
            "new_model_features_or_penalties": False,
            "reranking_remains_no_persist_experiment_only": True,
        },
        "summary": summary,
        "case_rows": all_case_rows,
        "governed_rerank_ready_rows": queues["governed_rerank_ready"],
        "first_pilot_case_rows": queues["governed_rerank_ready"],
        "baseline_support_rows": queues["baseline_support_only"],
        "spot_check_rows": queues["spot_check_only"],
        "analyst_review_rows": queues["analyst_review_only"],
        "hold_out_rows": queues["hold_out"],
        "fallback_safety_blocked_rows": queues["fallback_safety_blocked"],
        "no_reduction_no_action_rows": queues["no_reduction_no_action"],
        "fallback_blocked_rows": fallback_blocked,
        "opinion_of_value_rows": opinion_rows,
        "pilot_summary_rows": pilot_summary_rows,
        "column_key_rows": COLUMN_KEY_ROWS,
    }

    ts = payload["artifact_contract"]["created_at"]
    paths = output_paths(output_dir, ts)
    output_dir.mkdir(parents=True, exist_ok=True)
    paths["json"].write_text(json.dumps(payload, indent=2, sort_keys=True))
    write_csv(paths["csv"], all_case_rows)
    write_csv(paths["column_key"], COLUMN_KEY_ROWS)
    write_csv(paths["pilot_summary"], pilot_summary_rows)
    write_csv(paths["opinion_of_value"], opinion_rows)
    write_csv(paths["governed_rerank_ready"], queues["governed_rerank_ready"])
    write_csv(paths["first_pilot"], queues["governed_rerank_ready"])
    write_csv(paths["first_pilot_comp_details"], first_comp)
    write_csv(paths["baseline_support"], queues["baseline_support_only"])
    write_csv(paths["baseline_support_comp_details"], baseline_comp)
    write_csv(paths["signoff_tracker"], signoff_rows, fieldnames=SIGNOFF_FIELDNAMES)
    write_csv(paths["comparison_grid"], comparison_grid_rows)
    write_csv(paths["changed_comps_review"], changed_comp_rows)
    write_csv(paths["spot_check"], queues["spot_check_only"])
    write_csv(paths["analyst_review"], queues["analyst_review_only"])
    write_csv(paths["hold_out"], queues["hold_out"])
    write_csv(paths["fallback_safety_blocked"], queues["fallback_safety_blocked"])
    write_csv(paths["no_reduction_no_action"], queues["no_reduction_no_action"])
    write_csv(paths["fallback_blocked"], fallback_blocked)
    write_csv(paths["excluded_comp_details"], excluded_comp)
    render_markdown(paths["markdown"], payload)
    workbook_written = maybe_write_workbook(
        paths["workbook"],
        {
            "README_Key": COLUMN_KEY_ROWS,
            "Pilot_Summary": pilot_summary_rows,
            "Opinion_Of_Value": opinion_rows,
            "Analyst_Signoff": signoff_rows,
            "Changed_Comps_Review": changed_comp_rows,
            "Subject_Comp_Grid": comparison_grid_rows,
            "Full_First_Pilot_Comp_Details": first_comp,
            "Baseline_Support_Only": queues["baseline_support_only"],
            "Baseline_Support_Comp_Details": baseline_comp,
            "Spot_Check_Appendix": queues["spot_check_only"],
            "Analyst_Review_Only": queues["analyst_review_only"],
            "Hold_Out": queues["hold_out"],
            "Fallback_Safety_Blocked": queues["fallback_safety_blocked"],
            "No_Reduction_No_Action": queues["no_reduction_no_action"],
            "Fallback_Blocked": fallback_blocked,
        },
    )
    payload["summary"]["workbook_output"] = {
        "xlsx_written": workbook_written,
        "path": str(paths["workbook"]) if workbook_written else None,
    }
    paths["json"].write_text(json.dumps(payload, indent=2, sort_keys=True))
    if not workbook_written:
        paths.pop("workbook", None)
    return payload, paths


def main() -> None:
    args = build_parser().parse_args()
    payload, paths = build_packet(
        complete_comp_evidence_artifact=args.complete_comp_evidence_artifact,
        governed_fallback_artifacts=args.governed_fallback_artifact,
        raw_artifacts=args.raw_artifact,
        database_url=args.database_url,
        requested_tax_year=args.requested_tax_year,
        output_dir=args.output_dir,
        timestamp=args.timestamp,
    )
    for path in paths.values():
        print(path)
    print(json.dumps(payload["summary"]["first_pilot_ready"], sort_keys=True))
    print(json.dumps(payload["summary"]["first_pilot_county_summary"], sort_keys=True))


if __name__ == "__main__":
    main()
